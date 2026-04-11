"""Parseo de reportes, resumen, límites de plan, OpenAI y persistencia de análisis."""
from __future__ import annotations

import io
import json
import math
import os
import re
import secrets
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from fastapi import HTTPException, UploadFile

from config import (
    ADMIN_EMAILS,
    ALLOWED_UPLOAD_EXTENSIONS,
    DB_PATH,
    DEFAULT_MODEL,
    EXCEL_MAX_SHEETS_PER_FILE,
    FREE_MAX_ANALYSES,
    FREE_MAX_DAYS,
    FREE_MAX_FILES_PER_ANALYSIS,
    FREE_REPORTS_PER_MONTH,
    MAX_UPLOAD_BYTES_PER_FILE,
    OPENAI_API_KEY,
    PRO_180_MAX_ANALYSES,
    PRO_180_MAX_DAYS,
    PRO_180_MAX_FILES,
    PRO_90_MAX_ANALYSES,
    PRO_90_MAX_DAYS,
    PRO_90_MAX_FILES,
    PRO_90_REPORTS_PER_MONTH,
    PRO_PLUS_MAX_ANALYSES,
    PRO_PLUS_REPORTS_PER_MONTH,
    UPLOAD_DIR,
)
from db import db
from plan_entitlements import get_effective_plan
from time_utils import now_iso

MIN_BUSINESS_CONTEXT_LEN = 15


def require_business_context(business_context: Optional[str]) -> str:
    """Contexto de negocio obligatorio para orientar la lectura (web y API)."""
    s = (business_context or "").strip()
    if len(s) < MIN_BUSINESS_CONTEXT_LEN:
        raise ValueError(
            "El contexto es obligatorio: indica qué quieres entender o priorizar con esta lectura "
            f"(al menos {MIN_BUSINESS_CONTEXT_LEN} caracteres)."
        )
    return s


DATE_ALIASES = [
    "stay_date", "date", "business_date", "checkin", "check_in", "arrival", "arrival_date",
    "checkout", "check_out", "departure", "departure_date", "booking_date", "created_at",
    "reservation_date", "travel_date", "occupancy_date", "fecha", "fecha_noche", "dia",
]
REVENUE_ALIASES = [
    "net_room_revenue", "room_revenue", "revenue", "revenue_net", "net_revenue", "total_revenue",
    "room_rev", "amount", "booking_value", "room_amount", "total_amount", "subtotal", "total",
    "ingresos", "ingreso", "ingreso_total", "revenue_total",
]
GROSS_ALIASES = ["gross_revenue", "gross_amount", "gross_booking_value", "sell_amount"]
COMM_ALIASES = ["commission", "commission_amount", "ota_commission", "channel_commission"]
PAYMENT_ALIASES = ["payment_cost", "card_fee", "payment_fee", "processing_fee"]
ROOM_NIGHTS_ALIASES = ["room_nights", "nights", "night_count", "roomnights", "noches", "vendidas"]
RESERVATION_ID_ALIASES = ["reservation_id", "booking_id", "confirmation", "conf_no", "res_id"]
CHANNEL_ALIASES = ["channel", "source", "booking_channel", "channel_name", "ota", "segmento_canal", "canal"]
STATUS_ALIASES = ["status", "reservation_status", "booking_status", "estatus"]
ROOMS_ALIASES = ["rooms", "room_count", "habitaciones"]
ADR_ALIASES = ["adr", "average_daily_rate", "tarifa_promedio", "adr_usd"]
TAX_ALIASES = ["tax", "taxes", "vat", "impuestos"]


ANALYSIS_SCHEMA = {
    "type": "object",
    "properties": {
        "resumen_ejecutivo": {"type": "string"},
        "metricas_clave": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "nombre": {"type": "string"},
                    "valor": {"type": "string"},
                    "lectura": {"type": "string"},
                },
                "required": ["nombre", "valor", "lectura"],
                "additionalProperties": False,
            },
        },
        "hallazgos_prioritarios": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "titulo": {"type": "string"},
                    "detalle": {"type": "string"},
                    "impacto": {"type": "string"},
                    "prioridad": {"type": "string"},
                },
                "required": ["titulo", "detalle", "impacto", "prioridad"],
                "additionalProperties": False,
            },
        },
        "oportunidades_directo_vs_ota": {
            "type": "array",
            "items": {"type": "string"},
        },
        "riesgos_detectados": {
            "type": "array",
            "items": {"type": "string"},
        },
        "recomendaciones_accionables": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "accion": {"type": "string"},
                    "por_que": {"type": "string"},
                    "urgencia": {"type": "string"},
                },
                "required": ["accion", "por_que", "urgencia"],
                "additionalProperties": False,
            },
        },
        "datos_faltantes": {
            "type": "array",
            "items": {"type": "string"},
        },
        "senal_de_upgrade": {
            "type": "object",
            "properties": {
                "deberia_hacer_upgrade": {"type": "boolean"},
                "motivo": {"type": "string"},
            },
            "required": ["deberia_hacer_upgrade", "motivo"],
            "additionalProperties": False,
        },
    },
    "required": [
        "resumen_ejecutivo",
        "metricas_clave",
        "hallazgos_prioritarios",
        "oportunidades_directo_vs_ota",
        "riesgos_detectados",
        "recomendaciones_accionables",
        "datos_faltantes",
        "senal_de_upgrade",
    ],
    "additionalProperties": False,
}


def normalize_col(col: Any) -> str:
    col = str(col).strip().lower()
    col = re.sub(r"[^a-z0-9]+", "_", col)
    return col.strip("_")


def _dedupe_column_names(names: List[str]) -> List[str]:
    """
    Evita nombres duplicados: con columnas homónimas, pandas devuelve un DataFrame en df[col]
    y el análisis falla (p. ej. 'DataFrame' object has no attribute 'dtype').
    """
    counts: Dict[str, int] = {}
    out: List[str] = []
    for raw in names:
        base = (raw or "").strip() or "columna"
        n = counts.get(base, 0) + 1
        counts[base] = n
        if n == 1:
            out.append(base)
        else:
            out.append(f"{base}_{n}")
    return out


def _as_single_series(val: Any) -> Optional[pd.Series]:
    """Si llega un DataFrame (homónimos no deduplicados en otro flujo), usa la primera columna."""
    if val is None:
        return None
    if isinstance(val, pd.DataFrame):
        if val.shape[1] < 1:
            return None
        s = val.iloc[:, 0]
        return s.copy() if hasattr(s, "copy") else s
    if isinstance(val, pd.Series):
        return val
    try:
        return pd.Series(val, dtype=object)
    except Exception:
        return None


def _series_dtype(s: Any) -> Any:
    ser = _as_single_series(s)
    if ser is None or len(ser) == 0:
        return object
    return ser.dtype


def _dataframe_string_cells_for_retry(df: pd.DataFrame) -> pd.DataFrame:
    """Reintento cuando Excel mezcla tipos: todo como texto legible para fechas y montos."""
    out = df.copy()
    out.columns = _dedupe_column_names([normalize_col(str(c)) for c in out.columns])
    for c in out.columns:
        out[c] = out[c].map(lambda x: "" if pd.isna(x) else str(x).strip())
    return out


def maybe_to_datetime(series: Any) -> Optional[pd.Series]:
    series = _as_single_series(series)
    if series is None or len(series) == 0:
        return None
    try:
        converted = _series_try_datetime(series)
        success_ratio = converted.notna().mean()
        if success_ratio >= 0.45:
            return converted
        if series.dtype == object:
            fixed = series.map(_string_spanish_month_to_ascii)
            converted2 = pd.to_datetime(fixed, errors="coerce", dayfirst=True)
            if converted2.notna().mean() >= 0.45:
                return converted2
    except Exception:
        return None
    return None


def find_col(columns: List[str], aliases: List[str]) -> Optional[str]:
    for alias in aliases:
        if alias in columns:
            return alias
    for col in columns:
        for alias in aliases:
            if alias in col:
                return col
    return None


def _has_date_header(columns: Any) -> bool:
    """True si las columnas parecen encabezado de datos (tienen Fecha/Día/Date)."""
    lower = [str(c).lower() for c in columns]
    return any("fecha" in c or "date" in c or "día" in c or "dia" in c for c in lower)


# Palabras típicas en exportaciones PMS / channel (puntuación de fila candidata a encabezado)
_HEADER_KEYWORD_GROUPS: Tuple[Tuple[Tuple[str, ...], float], ...] = (
    (
        (
            "fecha",
            "date",
            "día",
            "dia",
            "arrival",
            "llegada",
            "departure",
            "salida",
            "checkin",
            "check-in",
            "checkout",
            "check-out",
            "stay",
            "created",
            "booking",
            "night",
            "noche",
        ),
        2.0,
    ),
    (
        (
            "revenue",
            "ingreso",
            "amount",
            "total",
            "tarifa",
            "rate",
            "adr",
            "importe",
            "valor",
            "usd",
            "mxn",
            "eur",
            "gross",
            "net",
            "room",
        ),
        1.5,
    ),
    (("channel", "canal", "source", "segment", "ota", "agency"), 1.2),
    (("commission", "comision", "comisión", "fee", "cargo", "payment"), 1.2),
    (("status", "estado", "estatus", "cancel", "void", "noshow"), 1.0),
    (("reserva", "reservation", "booking_id", "confirmation", "guest", "huésped", "huesped"), 1.0),
    (("rooms", "habitacion", "habitación", "room_nights", "noches"), 1.0),
)


def _score_header_labels(labels: List[Any]) -> float:
    """Mayor puntuación = más probable que la fila sea cabecera de tabla de reservas/ingresos."""
    score = 0.0
    non_empty = 0
    for x in labels:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            continue
        t = str(x).strip().lower()
        if not t or t == "nan":
            continue
        non_empty += 1
        if "unnamed" in t:
            score -= 0.6
            continue
        for kws, w in _HEADER_KEYWORD_GROUPS:
            if any(kw in t for kw in kws):
                score += w
    if non_empty < 2:
        score -= 40.0
    return score


def _tabular_bonus_next_row(row: pd.Series) -> float:
    """Bonus si la fila siguiente parece datos (números, importes, fechas parseables)."""
    if row is None or len(row) < 2:
        return 0.0
    ok = 0
    tot = 0
    for j in range(min(len(row), 28)):
        v = row.iloc[j]
        if pd.isna(v):
            continue
        tot += 1
        if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
            ok += 1
            continue
        if isinstance(v, str):
            t = v.strip().replace(" ", "").replace("$", "").replace(",", ".")
            if t and pd.notna(pd.to_numeric(t, errors="coerce")):
                ok += 1
            continue
        if pd.notna(pd.to_datetime(v, errors="coerce")):
            ok += 1
    if tot == 0:
        return 0.0
    r = ok / tot
    if r >= 0.42:
        return 3.0
    if r >= 0.22:
        return 1.4
    return 0.0


def _pick_best_header_row_index_with_baseline(
    preview: pd.DataFrame, max_scan: int = 22
) -> Tuple[int, float, float]:
    """Elige índice de fila (0-based) con mejor score; devuelve también score de la fila 0."""
    best_i = 0
    best_s = -1e9
    s0 = -1e9
    n = min(max_scan, len(preview))
    if n <= 0:
        return 0, -1e9, -1e9
    for i in range(n):
        labels = preview.iloc[i].tolist()
        s = _score_header_labels(labels)
        if i + 1 < len(preview):
            try:
                s += _tabular_bonus_next_row(preview.iloc[i + 1])
            except Exception:
                pass
        if i == 0:
            s0 = s
        if s > best_s:
            best_s = s
            best_i = i
    return best_i, best_s, s0


def _should_use_detected_header_row(best_i: int, best_s: float, s0: float) -> bool:
    if best_i <= 0:
        return False
    if best_s >= s0 + 1.25:
        return True
    if s0 < 2.5 and best_s >= 3.0:
        return True
    return False


def _sniff_csv_separator(raw: bytes, max_bytes: int = 16384) -> str:
    """Elige `,` o `;` según frecuencia en las primeras líneas (la fila 1 puede ser título sin separadores)."""
    head = raw[: min(max_bytes, len(raw))].decode("utf-8", errors="replace")
    lines = [ln.strip() for ln in head.splitlines()[:24] if ln.strip()]
    if not lines:
        return ","
    comma = sum(ln.count(",") for ln in lines)
    semi = sum(ln.count(";") for ln in lines)
    if semi > comma:
        return ";"
    return ","


def _csv_loose_preview_matrix(raw: bytes, sep: str, encoding: str, max_lines: int = 50) -> pd.DataFrame:
    """
    Matriz de celdas para puntuar cabeceras sin usar read_csv en bloque (filas previas al
    separador suelen tener menos columnas y pandas descarta líneas con on_bad_lines).
    """
    text = raw.decode(encoding, errors="replace").splitlines()[:max_lines]
    rows: List[List[str]] = []
    for ln in text:
        if not ln.strip():
            rows.append([""])
            continue
        rows.append([c.strip() for c in ln.split(sep)])
    if not rows:
        return pd.DataFrame()
    ncol = max(len(r) for r in rows)
    padded = [r + [""] * (ncol - len(r)) for r in rows]
    return pd.DataFrame(padded, dtype=object)


def _parse_csv_bytes_with_smart_header(raw: bytes, sep: str, filename: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    """Lee CSV eligiendo la fila de cabecera con mayor score (títulos / metadatos arriba de la tabla)."""
    notices: List[Dict[str, Any]] = []
    encodings = ("utf-8-sig", "utf-8", "latin-1")
    preview: Optional[pd.DataFrame] = None
    enc_ok = "utf-8"
    for enc in encodings:
        try:
            preview = _csv_loose_preview_matrix(raw, sep, enc)
            enc_ok = enc
            break
        except Exception:
            continue
    read_kw: Dict[str, Any] = {"sep": sep, "on_bad_lines": "skip"}
    if preview is None:
        for enc in encodings:
            try:
                return pd.read_csv(io.BytesIO(raw), encoding=enc, **read_kw), notices
            except Exception:
                continue
        raise ValueError("No se pudo leer el CSV (encoding o formato).")

    if preview.shape[1] < 2:
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc_ok, **read_kw), notices
        except Exception:
            for enc in encodings:
                try:
                    return pd.read_csv(io.BytesIO(raw), encoding=enc, **read_kw), notices
                except Exception:
                    continue
        raise ValueError("No se pudo leer el CSV.")

    read_kw["encoding"] = enc_ok
    best_i, best_s, s0 = _pick_best_header_row_index_with_baseline(preview)
    if _should_use_detected_header_row(best_i, best_s, s0):
        df = pd.read_csv(io.BytesIO(raw), header=0, skiprows=best_i, **read_kw)
        notices.append(
            {
                "kind": "header_auto_detected",
                "filename": filename,
                "message": (
                    f"«{filename}»: tomamos la fila {best_i + 1} como cabecera de la tabla "
                    f"(había texto o metadatos del export antes de los datos)."
                ),
            }
        )
        return df, notices
    return pd.read_csv(io.BytesIO(raw), header=0, **read_kw), notices


def _excel_apply_smart_header_row(
    raw: bytes,
    sheet: str,
    df_default: pd.DataFrame,
    mat_header_none: pd.DataFrame,
    filename: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    """Si la primera fila no es la cabecera real, re-parsea con header=best_i."""
    notices: List[Dict[str, Any]] = []
    if mat_header_none is None or mat_header_none.empty or mat_header_none.shape[1] < 2:
        return df_default, notices
    preview = mat_header_none.head(45)
    best_i, best_s, s0 = _pick_best_header_row_index_with_baseline(preview)
    if not _should_use_detected_header_row(best_i, best_s, s0):
        return df_default, notices
    try:
        kw: Dict[str, Any] = {}
        ext = os.path.splitext(filename or "")[1].lower()
        if ext == ".xls":
            kw = {"engine": "xlrd", "engine_kwargs": {"ignore_workbook_corruption": True}}
        df2 = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=best_i, **kw)
        if len(df2.columns) < 2:
            return df_default, notices
        notices.append(
            {
                "kind": "header_auto_detected",
                "filename": filename,
                "sheet": sheet,
                "message": (
                    f"«{filename}» (hoja «{sheet}»): cabecera de tabla en la fila {best_i + 1} "
                    f"respecto al archivo; las filas anteriores se trataron como título o notas."
                ),
            }
        )
        return df2, notices
    except Exception:
        return df_default, notices


# Meses abreviados en español (exports tipo 09/Abr/2026)
_SPANISH_MONTH_ABBR = (
    ("ene", "Jan"),
    ("feb", "Feb"),
    ("mar", "Mar"),
    ("abr", "Apr"),
    ("may", "May"),
    ("jun", "Jun"),
    ("jul", "Jul"),
    ("ago", "Aug"),
    ("sep", "Sep"),
    ("set", "Sep"),
    ("oct", "Oct"),
    ("nov", "Nov"),
    ("dic", "Dec"),
)

MES_ES_A_NUM = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def _string_spanish_month_to_ascii(val: Any) -> Any:
    if not isinstance(val, str):
        return val
    s = val.strip()
    if not s or s.lower() == "nan":
        return val
    out = s
    for es, en in _SPANISH_MONTH_ABBR:
        out = re.sub(rf"/{es}\b", f"/{en}", out, flags=re.IGNORECASE)
        out = re.sub(rf"-{es}\b", f"-{en}", out, flags=re.IGNORECASE)
        out = re.sub(rf"\b{es}\.", f"{en}.", out, flags=re.IGNORECASE)
    return out


def _series_try_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def _datelike_ratio(values: Any) -> float:
    """Ratio de celdas no nulas que pandas interpreta como fecha (muestra hasta 24 celdas)."""
    if values is None:
        return 0.0
    if hasattr(values, "iloc"):
        seq = [values.iloc[j] for j in range(min(len(values), 24))]
    else:
        seq = list(values)[:24]
    non_null = [v for v in seq if pd.notna(v)]
    if not non_null:
        return 0.0
    ok = 0
    for v in non_null:
        dt = pd.to_datetime(v, errors="coerce")
        if pd.isna(dt) and isinstance(v, str):
            dt = pd.to_datetime(_string_spanish_month_to_ascii(v), errors="coerce", dayfirst=True)
        if pd.notna(dt):
            ok += 1
    return ok / len(non_null)


def _wide_dates_metrics_to_long(mat: pd.DataFrame, header_row: int) -> Optional[pd.DataFrame]:
    """Primera columna = métrica; demás columnas = fechas (resumen producción por día)."""
    hdr = mat.iloc[header_row]
    date_idx: List[Tuple[int, pd.Timestamp]] = []
    for j in range(1, mat.shape[1]):
        v = hdr.iloc[j]
        if pd.isna(v):
            continue
        dt = pd.to_datetime(v, errors="coerce")
        if pd.isna(dt) and isinstance(v, str):
            dt = pd.to_datetime(_string_spanish_month_to_ascii(v), errors="coerce", dayfirst=True)
        if pd.notna(dt):
            date_idx.append((j, pd.Timestamp(dt).normalize()))
    if len(date_idx) < 3:
        return None
    records: List[Dict[str, Any]] = []
    for r in range(header_row + 1, len(mat)):
        label = mat.iloc[r, 0]
        if pd.isna(label):
            continue
        mname = str(label).strip()
        for j, dt in date_idx:
            val = mat.iloc[r, j]
            records.append({"fecha": dt, "metrica": mname, "valor": val})
    if not records:
        return None
    long_df = pd.DataFrame(records)
    wide = long_df.pivot_table(index="fecha", columns="metrica", values="valor", aggfunc="first")
    wide = wide.reset_index()
    wide.columns = _dedupe_column_names([normalize_col(str(c)) for c in wide.columns])
    return wide


def _ocupacion_mes_to_tidy(mat: pd.DataFrame, header_row: int) -> Optional[pd.DataFrame]:
    """Bloques 'YYYY - YYYY' + ADR + Revenue y filas por nombre de mes."""
    hdr = mat.iloc[header_row]
    groups: List[Tuple[int, int]] = []
    j = 0
    while j < len(hdr):
        v = hdr.iloc[j]
        if pd.isna(v):
            j += 1
            continue
        m = re.match(r"(\d{4})\s*-\s*\1", str(v).strip())
        if m:
            year = int(m.group(1))
            groups.append((year, j))
            j += 3
            continue
        j += 1
    if not groups:
        return None
    records: List[Dict[str, Any]] = []
    for r in range(header_row + 1, len(mat)):
        if mat.iloc[r].notna().sum() < 2:
            continue
        mes_cell = mat.iloc[r, 0]
        if pd.isna(mes_cell):
            continue
        mi = MES_ES_A_NUM.get(str(mes_cell).strip().lower())
        if not mi:
            continue
        for year, start_col in groups:
            if start_col + 2 >= mat.shape[1]:
                continue
            occ = mat.iloc[r, start_col]
            adr = mat.iloc[r, start_col + 1]
            rev = mat.iloc[r, start_col + 2]
            fecha = pd.Timestamp(year=year, month=mi, day=1)
            records.append(
                {
                    "fecha": fecha,
                    "ocupacion_pct": occ,
                    "adr": adr,
                    "ingreso": rev,
                }
            )
    if not records:
        return None
    return pd.DataFrame(records)


def _normalize_excel_matrix(mat: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Detecta layouts típicos de PMS (.xls) con títulos/impresión antes de la tabla real.
    Devuelve un DataFrame tabular o None si no aplica.
    """
    if mat is None or mat.empty:
        return None
    mat = mat.copy()
    nscan = min(25, len(mat))
    for i in range(nscan):
        row = mat.iloc[i]
        if len(row) == 0:
            continue
        s0 = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ""
        joined = " ".join(str(row.iloc[j]).lower() for j in range(min(len(row), 20)) if pd.notna(row.iloc[j]))

        if s0 == "canal" and len(row) > 1:
            h1 = str(row.iloc[1]).strip().lower() if pd.notna(row.iloc[1]) else ""
            if "llegada" in h1 or h1 == "arrivals":
                data = mat.iloc[i + 1 :].copy()
                hdr_cells = [str(x).strip() if pd.notna(x) else f"col_{k}" for k, x in enumerate(row)]
                data.columns = hdr_cells[: data.shape[1]]
                data = data.dropna(how="all").reset_index(drop=True)
                if len(data) > 0:
                    return data

        if "canal de venta" in s0:
            rest = row.iloc[1 : min(40, len(row))]
            if _datelike_ratio(rest) >= 0.5:
                wide = _wide_dates_metrics_to_long(mat, i)
                if wide is not None and len(wide) > 0:
                    return wide

        if "adr usd" in joined and "revenue usd" in joined:
            tidy = _ocupacion_mes_to_tidy(mat, i)
            if tidy is not None and len(tidy) > 0:
                return tidy

    return None


def _excel_to_dataframes_openpyxl(
    raw: bytes,
    filename: str,
    read_only: bool = True,
    extra_notices: Optional[List[Dict[str, Any]]] = None,
) -> List[Tuple[str, pd.DataFrame]]:
    """Fallback: leer Excel con openpyxl directo (soporta formatos que pandas ExcelFile a veces rechaza)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=read_only, data_only=True)
    result = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                mat = pd.DataFrame(rows)
                norm = _normalize_excel_matrix(mat)
                if norm is not None and len(norm) > 0:
                    result.append((f"{filename}::{sheet_name}", norm))
                    continue
                preview = mat.head(45)
                best_i, best_s, s0 = _pick_best_header_row_index_with_baseline(preview)
                if _should_use_detected_header_row(best_i, best_s, s0) and best_i < len(rows):
                    header = [str(c) if c is not None else "" for c in rows[best_i]]
                    data = rows[best_i + 1 :]
                    if extra_notices is not None:
                        extra_notices.append(
                            {
                                "kind": "header_auto_detected",
                                "filename": filename,
                                "sheet": sheet_name,
                                "message": (
                                    f"«{filename}» (hoja «{sheet_name}»): cabecera en la fila {best_i + 1} "
                                    f"(openpyxl); omitimos líneas previas al bloque de datos."
                                ),
                            }
                        )
                else:
                    header = [str(c) if c is not None else "" for c in rows[0]]
                    data = rows[1:]
                df = pd.DataFrame(data, columns=header)
                df = df.dropna(how="all").reset_index(drop=True)
            else:
                df = pd.DataFrame()
            result.append((f"{filename}::{sheet_name}", df))
    finally:
        if read_only:
            wb.close()
    return result


def _apply_excel_sheet_cap(
    filename: str, sheets: List[Tuple[str, pd.DataFrame]]
) -> Tuple[List[Tuple[str, pd.DataFrame]], List[Dict[str, Any]]]:
    """Acota cuántas hojas de un mismo libro Excel entran al análisis (transparencia operativa)."""
    cap = EXCEL_MAX_SHEETS_PER_FILE
    info: List[Dict[str, Any]] = []
    if cap <= 0 and len(sheets) > 1:
        info.append(
            {
                "kind": "excel_multi_sheet_note",
                "filename": filename,
                "sheets_in_file": len(sheets),
                "message": (
                    f"«{filename}» tiene {len(sheets)} pestañas; con la configuración actual "
                    "todas se leyeron y entran en el cruce del análisis."
                ),
            }
        )
    if cap <= 0 or len(sheets) <= cap:
        return sheets, info
    trimmed = sheets[:cap]
    omitted = len(sheets) - cap
    notice = {
        "kind": "excel_sheets_limited",
        "filename": filename,
        "sheets_read": cap,
        "sheets_in_file": len(sheets),
        "sheets_omitted": omitted,
        "message": (
            f"«{filename}» tiene {len(sheets)} pestañas; solo analizamos las primeras {cap} en esta corrida "
            f"({omitted} omitida(s)). No alteramos tu archivo: acotamos qué hojas entran al modelo. "
            "Para el resto, exporta otra pestaña como archivo aparte o haz otra lectura."
        ),
    }
    return trimmed, info + [notice]


def parse_file_with_notices(upload: UploadFile) -> Tuple[List[Tuple[str, pd.DataFrame]], List[Dict[str, Any]]]:
    notices: List[Dict[str, Any]] = []
    ext = os.path.splitext(upload.filename or "")[1].lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise ValueError(f"Formato no permitido. Solo: {', '.join(sorted(ALLOWED_UPLOAD_EXTENSIONS))}")
    raw = upload.file.read()
    if len(raw) > MAX_UPLOAD_BYTES_PER_FILE:
        raise ValueError(f"Archivo demasiado grande. Máximo {MAX_UPLOAD_BYTES_PER_FILE // (1024*1024)} MB por archivo.")
    if ext == ".csv":
        sep = _sniff_csv_separator(raw)
        fname = upload.filename or "reporte.csv"
        try:
            df, csv_notes = _parse_csv_bytes_with_smart_header(raw, sep, fname)
            notices.extend(csv_notes)
        except Exception:
            try:
                df = pd.read_csv(io.BytesIO(raw), sep=sep, encoding="utf-8", on_bad_lines="skip")
            except Exception:
                df = pd.read_csv(io.BytesIO(raw), encoding="utf-8", on_bad_lines="skip")
            if not _has_date_header(df.columns) and len(df) > 0:
                for skip in range(1, min(15, len(raw) // 60 + 1)):
                    try:
                        df2 = pd.read_csv(
                            io.BytesIO(raw), sep=sep, encoding="utf-8", skiprows=skip, on_bad_lines="skip"
                        )
                        if len(df2.columns) >= 2 and _has_date_header(df2.columns):
                            df = df2
                            notices.append(
                                {
                                    "kind": "header_auto_detected",
                                    "filename": fname,
                                    "message": (
                                        f"«{fname}»: cabecera con fecha detectada tras omitir {skip} fila(s) iniciales."
                                    ),
                                }
                            )
                            break
                    except Exception:
                        continue
        return [(fname, df)], notices
    if ext == ".xls":
        # Excel 97-2003: xlrd a veces marca como "corruptos" archivos válidos de PMS; ignorar esa comprobación
        _xls_kw = {"engine": "xlrd", "engine_kwargs": {"ignore_workbook_corruption": True}}
        try:
            excel = pd.ExcelFile(io.BytesIO(raw), **_xls_kw)
            result = []
            for sheet in excel.sheet_names:
                mat = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=None, **_xls_kw)
                norm = _normalize_excel_matrix(mat)
                if norm is not None and len(norm) > 0:
                    result.append((f"{upload.filename}::{sheet}", norm))
                    continue
                df = excel.parse(sheet)
                df, sn = _excel_apply_smart_header_row(
                    raw, sheet, df, mat, upload.filename or "archivo.xls"
                )
                notices.extend(sn)
                if not _has_date_header(df.columns) and len(df) >= 4:
                    for header_row in range(1, 15):
                        try:
                            df2 = excel.parse(sheet, header=header_row)
                            if len(df2.columns) >= 2 and _has_date_header(df2.columns):
                                df = df2
                                break
                        except Exception:
                            continue
                result.append((f"{upload.filename}::{sheet}", df))
            result, cap_notes = _apply_excel_sheet_cap(upload.filename or "archivo.xls", result)
            notices.extend(cap_notes)
            return result, notices
        except Exception as e:
            raise ValueError(
                "No pudimos leer este archivo Excel 97-2003 (.xls). "
                "Prueba exportando el reporte en CSV o en formato .xlsx."
            ) from e
    if ext in [".xlsx", ".xlsm"]:
        # Intentar primero con pandas (rápido y estándar)
        try:
            excel = pd.ExcelFile(io.BytesIO(raw))
            result = []
            for sheet in excel.sheet_names:
                mat = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=None)
                norm = _normalize_excel_matrix(mat)
                if norm is not None and len(norm) > 0:
                    result.append((f"{upload.filename}::{sheet}", norm))
                    continue
                df = excel.parse(sheet)
                df, sn = _excel_apply_smart_header_row(
                    raw, sheet, df, mat, upload.filename or "archivo.xlsx"
                )
                notices.extend(sn)
                if not _has_date_header(df.columns) and len(df) >= 4:
                    for header_row in range(1, 15):
                        try:
                            df2 = excel.parse(sheet, header=header_row)
                            if len(df2.columns) >= 2 and _has_date_header(df2.columns):
                                df = df2
                                break
                        except Exception:
                            continue
                result.append((f"{upload.filename}::{sheet}", df))
            result, cap_notes = _apply_excel_sheet_cap(upload.filename or "archivo.xlsx", result)
            notices.extend(cap_notes)
            return result, notices
        except Exception as e:
            err = str(e).strip().lower()
            # Si falla por “corruption” o formato raro, intentar con openpyxl en modo read_only (otra ruta de lectura)
            if "corruption" in err or "workbook" in err or "seen[" in err or "bad zip" in err or "invalid" in err:
                for use_read_only in (True, False):
                    try:
                        sheets = _excel_to_dataframes_openpyxl(
                            raw,
                            upload.filename or "reporte.xlsx",
                            read_only=use_read_only,
                            extra_notices=notices,
                        )
                        capped, cap_notes = _apply_excel_sheet_cap(
                            upload.filename or "archivo.xlsx", sheets
                        )
                        return capped, cap_notes
                    except Exception:
                        continue
                raise ValueError(
                    "No pudimos leer este Excel (formato propio de tu PMS/channel). "
                    "Estamos mejorando el soporte. Mientras tanto: exporta el mismo reporte en CSV y súbelo aquí."
                ) from e
            raise
    raise ValueError(f"Formato no soportado: {ext}")


def parse_file(upload: UploadFile) -> List[Tuple[str, pd.DataFrame]]:
    sheets, _ = parse_file_with_notices(upload)
    return sheets


def _infer_sheet_compute(sheet_name: str, df: pd.DataFrame) -> Dict[str, Any]:
    """Inferencia principal; errores parciales se omiten con avisos en sheet_warnings."""
    warnings: List[str] = []
    original_cols = [str(c) for c in df.columns]
    norm_cols = _dedupe_column_names([normalize_col(c) for c in original_cols])
    work = df.copy()
    work.columns = norm_cols
    work = work.dropna(how="all")
    if len(work) > 5000:
        work = work.head(5000)

    mappings = {
        "reservation_id": find_col(norm_cols, RESERVATION_ID_ALIASES),
        "channel": find_col(norm_cols, CHANNEL_ALIASES),
        "status": find_col(norm_cols, STATUS_ALIASES),
        "revenue": find_col(norm_cols, REVENUE_ALIASES),
        "gross_revenue": find_col(norm_cols, GROSS_ALIASES),
        "commission": find_col(norm_cols, COMM_ALIASES),
        "payment_cost": find_col(norm_cols, PAYMENT_ALIASES),
        "room_nights": find_col(norm_cols, ROOM_NIGHTS_ALIASES),
        "rooms": find_col(norm_cols, ROOMS_ALIASES),
        "adr": find_col(norm_cols, ADR_ALIASES),
        "taxes": find_col(norm_cols, TAX_ALIASES),
    }

    date_candidates: List[str] = []
    for col in norm_cols:
        if any(alias in col for alias in DATE_ALIASES):
            try:
                converted = maybe_to_datetime(work[col])
                if converted is not None:
                    work[col] = converted
                    date_candidates.append(col)
            except Exception as e:
                warnings.append(f"Columna «{col}» (fecha por nombre): omitida ({type(e).__name__}).")

    for col in norm_cols:
        if col not in date_candidates and _series_dtype(work[col]) == object:
            try:
                converted = maybe_to_datetime(work[col])
                if converted is not None and len(date_candidates) < 4:
                    work[col] = converted
                    date_candidates.append(col)
            except Exception as e:
                warnings.append(f"Columna «{col}» (fecha heurística): omitida ({type(e).__name__}).")

    checkin_col = next((c for c in date_candidates if any(x in c for x in ["checkin", "check_in", "arrival"])), None)
    checkout_col = next((c for c in date_candidates if any(x in c for x in ["checkout", "check_out", "departure"])), None)
    stay_col = next((c for c in date_candidates if any(x in c for x in ["stay", "business_date", "occupancy"])), None) or checkin_col
    booking_col = next((c for c in date_candidates if any(x in c for x in ["booking", "created", "reservation_date"])), None)

    date_col_set = set(date_candidates)
    for col in norm_cols:
        if col in date_col_set:
            continue
        try:
            s = work[col]
            if pd.api.types.is_datetime64_any_dtype(s):
                continue
            converted = pd.to_numeric(s, errors="coerce")
            if converted.notna().mean() >= 0.5:
                work[col] = converted
        except Exception:
            pass

    row_count = int(len(work))
    rev_col = mappings["revenue"] or mappings["gross_revenue"]
    comm_col = mappings["commission"]
    pay_col = mappings["payment_cost"]
    rn_col = mappings["room_nights"]
    channel_col = mappings["channel"]
    status_col = mappings["status"]
    adr_col = mappings["adr"]

    def _to_numeric_currency(series: Any) -> pd.Series:
        s = _as_single_series(series)
        if s is None or len(s) == 0:
            return pd.Series(dtype=float)
        if s.dtype != object and pd.api.types.is_numeric_dtype(s):
            return s
        cleaned = s.astype(str).str.replace(r"[$,\s]", "", regex=True)
        return pd.to_numeric(cleaned, errors="coerce")

    for col in [rev_col, comm_col, pay_col, adr_col]:
        if col and col in work.columns:
            try:
                work[col] = _to_numeric_currency(work[col])
            except Exception as e:
                warnings.append(f"Columna «{col}» (moneda): sin convertir ({type(e).__name__}).")

    metrics: Dict[str, Any] = {}
    if rev_col:
        try:
            metrics["revenue_total"] = round(float(work[rev_col].fillna(0).sum()), 2)
        except Exception as e:
            warnings.append(f"Ingreso total: no calculado ({type(e).__name__}).")
    if comm_col:
        try:
            metrics["comision_total"] = round(float(work[comm_col].fillna(0).sum()), 2)
        except Exception as e:
            warnings.append(f"Comisiones: no calculadas ({type(e).__name__}).")
    if pay_col:
        try:
            metrics["costo_pago_total"] = round(float(work[pay_col].fillna(0).sum()), 2)
        except Exception as e:
            warnings.append(f"Costo de pago: no calculado ({type(e).__name__}).")
    if rn_col:
        try:
            metrics["room_nights"] = round(float(work[rn_col].fillna(0).sum()), 2)
        except Exception as e:
            warnings.append(f"Room nights: no calculadas ({type(e).__name__}).")
    elif checkin_col and checkout_col:
        try:
            metrics["room_nights"] = round(
                float((work[checkout_col] - work[checkin_col]).dt.days.clip(lower=0).fillna(0).sum()), 2
            )
        except Exception as e:
            warnings.append(f"Noches por fechas entrada/salida: no calculadas ({type(e).__name__}).")
    if adr_col:
        try:
            metrics["adr_promedio"] = round(float(pd.to_numeric(work[adr_col], errors="coerce").mean()), 2)
        except Exception as e:
            warnings.append(f"ADR: no calculado ({type(e).__name__}).")
    elif metrics.get("revenue_total") and metrics.get("room_nights"):
        if metrics["room_nights"] > 0:
            metrics["adr_estimado"] = round(metrics["revenue_total"] / metrics["room_nights"], 2)

    if status_col:
        try:
            s = work[status_col].astype(str).str.lower()
            cancelled = int(s.str.contains("cancel|void|no show|noshow").sum())
            metrics["cancelaciones"] = cancelled
            metrics["cancelacion_pct"] = round((cancelled / max(row_count, 1)) * 100, 2)
        except Exception as e:
            warnings.append(f"Cancelaciones: no calculadas ({type(e).__name__}).")

    if channel_col:
        try:
            channel_series = work[channel_col].astype(str).str.strip().replace("", "Sin canal")
            top_channels = channel_series.value_counts().head(8)
            metrics["top_canales_por_reservas"] = [
                {"canal": str(k), "reservas": int(v)} for k, v in top_channels.items()
            ]
        except Exception as e:
            warnings.append(f"Ranking de canales (reservas): omitido ({type(e).__name__}).")
        try:
            if channel_col and rev_col:
                channel_series = work[channel_col].astype(str).str.strip().replace("", "Sin canal")
                rev_by_channel = work.groupby(channel_series)[rev_col].sum().sort_values(ascending=False).head(8)
                metrics["top_canales_por_ingreso"] = [
                    {"canal": str(k), "ingreso": round(float(v), 2)} for k, v in rev_by_channel.items()
                ]
        except Exception as e:
            warnings.append(f"Ranking de canales (ingreso): omitido ({type(e).__name__}).")
        try:
            if channel_col and comm_col and rev_col:
                channel_series = work[channel_col].astype(str).str.strip().replace("", "Sin canal")
                grouped = work.groupby(channel_series)[[rev_col, comm_col]].sum()
                channel_margin = []
                for chan, row in grouped.iterrows():
                    margin = float(row[rev_col]) - float(row[comm_col])
                    channel_margin.append(
                        {
                            "canal": str(chan),
                            "ingreso": round(float(row[rev_col]), 2),
                            "comision": round(float(row[comm_col]), 2),
                            "margen_estimado": round(margin, 2),
                        }
                    )
                metrics["margen_estimado_por_canal"] = sorted(
                    channel_margin, key=lambda x: x["margen_estimado"], reverse=True
                )[:8]
        except Exception as e:
            warnings.append(f"Margen por canal: omitido ({type(e).__name__}).")

    min_date = None
    max_date = None
    for col in [stay_col, checkin_col, booking_col, checkout_col]:
        if col and col in work.columns:
            try:
                col_min = work[col].dropna().min()
                col_max = work[col].dropna().max()
                if col_min is not pd.NaT and col_max is not pd.NaT:
                    if min_date is None or col_min < min_date:
                        min_date = col_min
                    if max_date is None or col_max > max_date:
                        max_date = col_max
            except Exception:
                pass
    if (min_date is None or max_date is None) and date_candidates:
        for col in date_candidates:
            if col not in work.columns:
                continue
            try:
                col_min = work[col].dropna().min()
                col_max = work[col].dropna().max()
                if col_min is not pd.NaT and col_max is not pd.NaT:
                    if min_date is None or col_min < min_date:
                        min_date = col_min
                    if max_date is None or col_max > max_date:
                        max_date = col_max
                    break
            except Exception:
                pass

    try:
        min_date_norm = pd.Timestamp(min_date) if min_date is not None else None
    except (TypeError, ValueError):
        min_date_norm = None
    try:
        max_date_norm = pd.Timestamp(max_date) if max_date is not None else None
    except (TypeError, ValueError):
        max_date_norm = None

    days_covered = 0
    if (
        min_date_norm is not None
        and max_date_norm is not None
        and pd.notna(min_date_norm)
        and pd.notna(max_date_norm)
    ):
        try:
            delta = max_date_norm - min_date_norm
            days_covered = int(getattr(delta, "days", delta)) + 1
        except (TypeError, ValueError):
            days_covered = 0

    fields_detected = [k for k, v in mappings.items() if v]
    out: Dict[str, Any] = {
        "sheet_name": sheet_name,
        "rows": row_count,
        "mappings": mappings,
        "date_columns": {
            "stay": stay_col,
            "booking": booking_col,
            "checkin": checkin_col,
            "checkout": checkout_col,
        },
        "fields_detected": fields_detected,
        "metrics": metrics,
        "days_covered": days_covered,
        "date_range": {
            "start": min_date_norm.isoformat()
            if min_date_norm is not None and not pd.isna(min_date_norm)
            else None,
            "end": max_date_norm.isoformat()
            if max_date_norm is not None and not pd.isna(max_date_norm)
            else None,
        },
        "sample_columns": norm_cols[:30],
    }
    if warnings:
        out["sheet_warnings"] = warnings
    return out


def _sheet_recovery_infer(sheet_name: str, df: pd.DataFrame, exc: BaseException) -> Dict[str, Any]:
    """
    Último nivel: no devolver vacío si aún podemos mapear columnas por nombre y sumar/fechar.
    Se usa tras fallo total de _infer_sheet_compute o reintento con celdas texto.
    """
    warnings: List[str] = [
        f"Lectura en modo recuperación ({type(exc).__name__}). Seguimos con lo que se pueda inferir del archivo."
    ]
    empty_mappings = {
        "reservation_id": None,
        "channel": None,
        "status": None,
        "revenue": None,
        "gross_revenue": None,
        "commission": None,
        "payment_cost": None,
        "room_nights": None,
        "rooms": None,
        "adr": None,
        "taxes": None,
    }
    if df is None or df.empty or len(df.columns) == 0:
        return {
            "sheet_name": sheet_name,
            "rows": int(len(df)) if df is not None else 0,
            "mappings": empty_mappings,
            "date_columns": {"stay": None, "booking": None, "checkin": None, "checkout": None},
            "fields_detected": [],
            "metrics": {},
            "days_covered": 0,
            "date_range": {"start": None, "end": None},
            "sample_columns": [],
            "sheet_warnings": warnings + [str(exc)],
        }

    try:
        work = df.copy()
        norm_cols = _dedupe_column_names([normalize_col(str(c)) for c in df.columns])
        work.columns = norm_cols
        work = work.dropna(how="all").head(5000)
    except Exception:
        return {
            "sheet_name": sheet_name,
            "rows": int(len(df)),
            "mappings": empty_mappings,
            "date_columns": {"stay": None, "booking": None, "checkin": None, "checkout": None},
            "fields_detected": [],
            "metrics": {},
            "days_covered": 0,
            "date_range": {"start": None, "end": None},
            "sample_columns": [],
            "sheet_warnings": warnings + [str(exc)],
        }

    mappings = {
        "reservation_id": find_col(norm_cols, RESERVATION_ID_ALIASES),
        "channel": find_col(norm_cols, CHANNEL_ALIASES),
        "status": find_col(norm_cols, STATUS_ALIASES),
        "revenue": find_col(norm_cols, REVENUE_ALIASES),
        "gross_revenue": find_col(norm_cols, GROSS_ALIASES),
        "commission": find_col(norm_cols, COMM_ALIASES),
        "payment_cost": find_col(norm_cols, PAYMENT_ALIASES),
        "room_nights": find_col(norm_cols, ROOM_NIGHTS_ALIASES),
        "rooms": find_col(norm_cols, ROOMS_ALIASES),
        "adr": find_col(norm_cols, ADR_ALIASES),
        "taxes": find_col(norm_cols, TAX_ALIASES),
    }
    metrics: Dict[str, Any] = {}
    rev_c = mappings["revenue"] or mappings["gross_revenue"]
    if rev_c:
        try:
            s = pd.to_numeric(_as_single_series(work[rev_c]), errors="coerce")
            if s is not None and s.notna().any():
                metrics["revenue_total"] = round(float(s.fillna(0).sum()), 2)
        except Exception:
            pass
    if mappings["commission"]:
        try:
            s = pd.to_numeric(_as_single_series(work[mappings["commission"]]), errors="coerce")
            if s is not None and s.notna().any():
                metrics["comision_total"] = round(float(s.fillna(0).sum()), 2)
        except Exception:
            pass
    if mappings["room_nights"]:
        try:
            s = pd.to_numeric(_as_single_series(work[mappings["room_nights"]]), errors="coerce")
            if s is not None and s.notna().any():
                metrics["room_nights"] = round(float(s.fillna(0).sum()), 2)
        except Exception:
            pass
    if mappings["channel"]:
        try:
            cs = work[mappings["channel"]].astype(str).str.strip().replace("", "Sin canal")
            vc = cs.value_counts().head(8)
            metrics["top_canales_por_reservas"] = [{"canal": str(k), "reservas": int(v)} for k, v in vc.items()]
        except Exception:
            pass

    stay_guess = None
    min_date_norm = None
    max_date_norm = None
    for col in norm_cols:
        try:
            m = maybe_to_datetime(work[col])
            if m is None or m.notna().sum() < max(2, int(len(work) * 0.2)):
                continue
            stay_guess = col
            dr = m.dropna()
            if len(dr) == 0:
                continue
            min_date_norm = pd.Timestamp(dr.min())
            max_date_norm = pd.Timestamp(dr.max())
            break
        except Exception:
            continue

    days_covered = 0
    if (
        min_date_norm is not None
        and max_date_norm is not None
        and pd.notna(min_date_norm)
        and pd.notna(max_date_norm)
    ):
        try:
            days_covered = int((max_date_norm - min_date_norm).days) + 1
        except (TypeError, ValueError):
            days_covered = 0

    fields_detected = [k for k, v in mappings.items() if v]
    return {
        "sheet_name": sheet_name,
        "rows": int(len(work)),
        "mappings": mappings,
        "date_columns": {
            "stay": stay_guess,
            "booking": None,
            "checkin": None,
            "checkout": None,
        },
        "fields_detected": fields_detected,
        "metrics": metrics,
        "days_covered": days_covered,
        "date_range": {
            "start": min_date_norm.isoformat() if min_date_norm is not None and pd.notna(min_date_norm) else None,
            "end": max_date_norm.isoformat() if max_date_norm is not None and pd.notna(max_date_norm) else None,
        },
        "sample_columns": norm_cols[:30],
        "sheet_warnings": warnings,
    }


def infer_sheet(sheet_name: str, df: pd.DataFrame, _recover_depth: int = 0) -> Dict[str, Any]:
    """
    Orquesta inferencia + reintento con celdas como texto + recuperación mínima con métricas útiles.
    Evita devolver resúmenes vacíos si el archivo aún permite sumar ingresos o acotar fechas.
    """
    if df is None:
        return _sheet_recovery_infer(sheet_name, pd.DataFrame(), ValueError("sin datos"))
    try:
        return _infer_sheet_compute(sheet_name, df)
    except Exception as e:
        if _recover_depth < 1:
            try:
                df_fix = _dataframe_string_cells_for_retry(df)
                out = infer_sheet(sheet_name, df_fix, _recover_depth=1)
                sw = out.setdefault("sheet_warnings", [])
                sw.insert(0, f"Reintento tras normalizar tipos de celda ({type(e).__name__}).")
                return out
            except Exception:
                pass
        return _sheet_recovery_infer(sheet_name, df, e)


def summarize_reports(files: List[UploadFile]) -> Dict[str, Any]:
    report_summaries = []
    max_days = 0
    all_starts = []
    all_ends = []
    upload_notices: List[Dict[str, Any]] = []
    for upload in files:
        sheets, notices = parse_file_with_notices(upload)
        upload_notices.extend(notices)
        for sheet_name, df in sheets:
            summary = infer_sheet(sheet_name, df)
            report_summaries.append(summary)
            sw = summary.get("sheet_warnings")
            if sw:
                upload_notices.append(
                    {
                        "kind": "sheet_infer_warnings",
                        "sheet": sheet_name,
                        "messages": sw,
                    }
                )
            max_days = max(max_days, summary.get("days_covered", 0))
            dr = summary.get("date_range", {})
            if dr.get("start"):
                t = pd.to_datetime(dr["start"])
                if pd.notna(t):
                    all_starts.append(t)
            if dr.get("end"):
                t = pd.to_datetime(dr["end"])
                if pd.notna(t):
                    all_ends.append(t)
    overall_days = 0
    if all_starts and all_ends:
        try:
            delta = max(all_ends) - min(all_starts)
            d = getattr(delta, "days", None)
            if d is not None and not (isinstance(d, float) and math.isnan(d)):
                overall_days = int(d) + 1
        except (TypeError, ValueError):
            pass
    return {
        "total_files": len(files),
        "reports_detected": len(report_summaries),
        "max_days_covered": max_days,
        "overall_days_covered": overall_days,
        "report_summaries": report_summaries,
        "upload_notices": upload_notices,
    }


HOTEL_PROMPT = """
Eres DRAGONNÉ, consultor senior en revenue management, distribución y e-commerce hotelero (cadenas y propiedades independientes en LATAM). Redactas como memorando comercial para dirección: profesional, claro, sobrio y accionable. Usa lenguaje de hotelería real. Evita tono de marketing, frases tipo “IA”, exageraciones, alarmismo innecesario y palabras rebuscadas.

Recibirás un JSON con:
- plan: "free_30", "pro_90" o "pro_180".
- contexto_hotel: objeto con hotel_nombre, hotel_tamano, hotel_categoria, hotel_ubicacion, hotel_estrellas, hotel_ubicacion_destino; además (opcionales): hotel_pms, hotel_channel_manager, hotel_booking_engine, hotel_tech_other (stack tecnológico), hotel_google_business_url, hotel_expedia_url, hotel_booking_url (enlaces de presencia online). Úsalo SIEMPRE para personalizar.
- contexto_negocio: qué le preocupa al usuario.
- resumen: lectura estructurada de reportes exportados desde PMS / channel / motor (CSV/Excel), incluyendo métricas por canal, días cubiertos y fechas.

Objetivo: que un gerente general o director comercial entienda la foto en pocos minutos y que un revenue manager tenga decisiones concretas sobre tarifa, canales y costo de distribución.

Reglas de estilo general:
1) No repitas el reporte fila por fila: explica QUÉ IMPLICA para ingresos, mezcla de canales, tarifa y margen después de comisiones. Cada cifra va con lectura breve (“esto sugiere…”, “conviene vigilar…”, “habría que contrastar con…”).
2) Tono de socio de negocio: directo y honesto. Si hay concentración de riesgo (dependencia de una OTA, tarifa baja en días fuertes, huecos entre semana), dilo con hechos; sin titulares sensacionalistas ni catastrofismo.
3) Estructura siempre la respuesta en bloques claros del JSON: resumen_ejecutivo, hallazgos_prioritarios, recomendaciones_accionables, datos_faltantes. Prioriza 3–7 puntos por bloque; evita listas largas sin jerarquía.

Reglas analíticas (cómo leer los datos):
4) Palancas que deben aparecer SIEMPRE que haya datos suficientes:
   - Mezcla de canales: canal directo vs OTAs, dependencia de 1–2 agencias, canales con mucho volumen pero ADR o margen neto claramente más bajo.
   - Pricing y ADR: diferencias de tarifa entre canales para las mismas fechas, entre semana vs fin de semana, y entre tipos de día (p.ej. noches con eventos/local demand drivers si las detectas).
   - Ritmo y distribución en el tiempo: qué días de la semana cargan la ocupación (¿solo viernes-sábado?, ¿se cae domingo-jueves?), tendencias dentro del rango (primer tercio vs último tercio).
   - Coste de distribución: cuando haya comisiones o % estimados, habla de margen neto y no solo de revenue bruto; si no hay comisiones, dilo y explica por qué sería clave tenerlas.
   - Riesgos y oportunidades de inventario: noches con poca demanda donde tiene sentido empujar volumen, y noches fuertes donde conviene proteger tarifa y mix directo.
5) Cuando veas patrones interesantes, dilo explícitamente. Ejemplos de redacción:
   - “Tu ocupación se concentra casi solo en viernes y sábado; de domingo a jueves el hotel funciona a media máquina.”
   - “Booking.com te aporta la mayor producción, pero a un ADR un 12–15% más bajo que el resto de tus canales en las mismas fechas; eso está erosionando margen.”
   - “El canal directo casi no participa en fines de semana fuertes; estás dejando toda la fecha en manos de OTAs con comisión.”

Uso del contexto del hotel:
6) CONTEXTO DEL HOTEL (obligatorio): El resumen_ejecutivo y las recomendaciones_accionables deben estar explícitamente adaptados al tipo de propiedad (hotel_tamano), categoría (hotel_categoria), estrellas (hotel_estrellas si > 0), ubicación geográfica (hotel_ubicacion) y posición en el destino (hotel_ubicacion_destino: centro, zona turística o periferia). No des el mismo consejo a un boutique que a un resort, ni a un hotel de ciudad que a un all-inclusive; adapta tono y expectativas a estrellas y ubicación (centro vs periferia).
7) Usa el stack tecnológico (PMS, channel manager, booking engine, otras tecnologías) para proponer acciones realistas: por ejemplo, hablar de reglas en el channel manager, ajustes de inventario hacia canal directo, paquetes en el motor de reservas, etc., siempre que tenga sentido.
8) IMPORTANTE sobre enlaces (Google My Business, Expedia, Booking): los URLs en contexto_hotel son SOLO para contexto (saber dónde está el hotel online). NUNCA inventes datos, reseñas ni métricas a partir de esas URLs; usa solo lo que viene en el resumen de reportes subidos.

Recomendaciones y enfoque accionable:
9) Recomendaciones: nada genérico tipo "revisar pricing". Sé específico en qué moverías (tarifa, canal, estancia mínima, política de cancelación, bundles, visibilidad por canal), en qué fechas / rangos de días / tipos de día y con qué objetivo (subir margen, ganar directo, reducir dependencia de OTA, etc.); y que sean coherentes con la categoría y tipo de propiedad del hotel.
10) Siempre que sea posible, traduce el hallazgo en un “plan de juego” para las próximas 2–4 semanas, agrupando acciones por prioridad: qué moverías primero, qué dejarías en observación y qué revisarías cuando haya más datos.

Datos faltantes y uso responsable:
11) No inventes datos; si asumes algo, dilo. Marca estimaciones como tales y mantén rangos conservadores.
12) Datos faltantes: en tono operativo, indica qué huecos tiene el export (canales, comisiones, noches vendidas, fechas de estancia, segmentos, etc.) y qué tipo de archivo convendría sumar. No hables del “modelo” ni uses tono disculpante: es información que falta en los datos cargados. No uses la falta de datos como excusa para no extraer valor de lo que sí existe.
13) REPORTES SIN CANALES (forecast, revenue por día): Si el resumen no incluye canales de distribución pero SÍ tiene ingresos, ADR y/o fechas (p. ej. reporte Forecast and Revenue con una fila por día), OBLIGATORIO extraer valor de lo que sí hay. Analiza: ingresos y ADR por día de la semana, diferencia entre semana vs fin de semana, tendencia en el mes, oportunidades de precio entre semana. Da recomendaciones concretas usando esos números. En resumen_ejecutivo incluye una línea tipo: "Esto es lo que podemos rescatar de tu reporte; con reportes que incluyan canales de distribución podemos aportar aún más valor (mix directo vs OTA, comisiones)." En datos_faltantes menciona canales/comisiones como mejora futura, no como excusa para no analizar lo que sí está.

Planes y señal de upgrade:
14) Planes:
   - "free_30": trata el caso como diagnóstico rápido sobre un máximo de 30 días; prioriza quick wins y, si faltan periodos para tener contexto, explícalo en la señal de upgrade.
   - "pro_90" y "pro_180": además de la foto actual, compara dentro del rango primeros días vs últimos (por ejemplo, primer tercio vs último tercio) y marca en hallazgos_prioritarios cualquier cambio de tendencia relevante (mix de canal, ADR, margen, cancelaciones). Usa también recomendaciones_accionables para proponer cómo reaccionar a esos cambios.
15) Señal de upgrade: si con más días o más reportes la lectura sería mucho más fuerte, dilo con tacto y con un ejemplo concreto (ej. "comparar este trimestre con el mismo periodo del año pasado").
16) Las integraciones directas con PMS (Mews, Cloudbeds, Little Hotelier) están pensadas como "Próximamente"; no prometas nada distinto a eso.

Devuelve SIEMPRE exclusivamente el JSON con el esquema especificado por el backend, sin texto adicional.
""".strip()


def call_openai(summary: Dict[str, Any], business_context: str, hotel_context: Dict[str, Any], plan: str) -> Dict[str, Any]:
    if not OPENAI_API_KEY:
        raise RuntimeError("No encontré OPENAI_API_KEY en el backend.")

    payload = {
        "model": DEFAULT_MODEL,
        "input": [
            {
                "role": "system",
                "content": [{"type": "input_text", "text": HOTEL_PROMPT}],
            },
            {
                "role": "user",
                "content": [{
                    "type": "input_text",
                    "text": json.dumps({
                        "plan": plan,
                        "contexto_hotel": hotel_context,
                        "uso_contexto_hotel": "Usa siempre contexto_hotel (nombre, tipo de propiedad/tamaño, categoría, ubicación) para redactar resumen_ejecutivo y recomendaciones_accionables de forma específica a este hotel.",
                        "contexto_negocio": business_context or "Sin notas de contexto de negocio del usuario.",
                        "resumen": summary,
                    }, ensure_ascii=False),
                }],
            },
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "profitpilot_hotel_analysis",
                "schema": ANALYSIS_SCHEMA,
                "strict": True,
            }
        },
    }
    response = requests.post(
        "https://api.openai.com/v1/responses",
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=120,
    )
    response.raise_for_status()
    data = response.json()
    text = None
    if data.get("output"):
        for item in data["output"]:
            for content in item.get("content", []):
                if content.get("type") in {"output_text", "text"} and content.get("text"):
                    text = content["text"]
                    break
            if text:
                break
    if not text:
        text = data.get("output_text") or data.get("text", "")
    if not text:
        raise RuntimeError("No pude leer la respuesta estructurada de OpenAI.")
    return json.loads(text)


def analyses_count(user_id: int) -> int:
    with db() as conn:
        row = conn.execute("SELECT COUNT(*) AS c FROM analyses WHERE user_id = ?", (user_id,)).fetchone()
        return int(row["c"])


def analyses_this_month(user_id: int) -> int:
    """Análisis aún presentes en BD creados en el mes actual (UTC). Preferir generations_this_month para cupos."""
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()[:19]
    with db() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM analyses WHERE user_id = ? AND created_at >= ?",
            (user_id, start_of_month),
        ).fetchone()
        return int(row["c"])


def generations_this_month(user_id: int) -> int:
    """Lecturas completadas en el mes UTC según registro append-only (no baja si se borra el análisis)."""
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    with db() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM analysis_run_log WHERE user_id = ? AND created_at >= ?",
            (user_id, start_of_month),
        ).fetchone()
        return int(row["c"])


def _monthly_generation_limit_for_plan(plan: str) -> int:
    if plan == "free":
        return FREE_REPORTS_PER_MONTH
    if plan == "pro":
        return PRO_90_REPORTS_PER_MONTH
    if plan in ("pro_plus", "free_trial"):
        return PRO_PLUS_REPORTS_PER_MONTH
    return FREE_REPORTS_PER_MONTH


def reserve_monthly_generation_or_raise(user_id: int, plan: str) -> int:
    """
    Reserva atómicamente un cupo mensual antes de llamar al modelo (evita carreras entre peticiones paralelas).
    Devuelve el id de fila en analysis_run_log (analysis_id NULL hasta guardar). Lanza 402 si el mes está lleno.
    """
    lim = _monthly_generation_limit_for_plan(plan)
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    created = now_iso()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("BEGIN IMMEDIATE")
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM analysis_run_log WHERE user_id = ? AND created_at >= ?",
            (user_id, start_of_month),
        ).fetchone()
        if int(row["c"]) >= lim:
            conn.rollback()
            if plan == "free":
                detail = "Ya usaste tu reporte gratuito de este mes. Pásate a Pro para seguir subiendo."
            elif plan == "pro":
                detail = (
                    f"Has alcanzado el límite de lecturas de este mes en Pro ({PRO_90_REPORTS_PER_MONTH}). "
                    "Renueva el cupo el próximo mes natural (UTC) o sube a Pro+."
                )
            else:
                detail = (
                    f"Has alcanzado el límite de lecturas de este mes en Pro+ ({PRO_PLUS_REPORTS_PER_MONTH}). "
                    "El cupo se renueva el próximo mes natural (UTC). Contáctanos si necesitas más capacidad."
                )
            raise HTTPException(status_code=402, detail=detail)
        cur = conn.execute(
            "INSERT INTO analysis_run_log (user_id, analysis_id, created_at) VALUES (?, NULL, ?)",
            (user_id, created),
        )
        rid = int(cur.lastrowid)
        conn.commit()
        return rid
    except HTTPException:
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def release_reserved_generation_row(reserved_log_id: int, user_id: int) -> None:
    """Libera una reserva si falló el modelo o el guardado antes de completar el análisis."""
    if not reserved_log_id:
        return
    with db() as conn:
        conn.execute(
            "DELETE FROM analysis_run_log WHERE id = ? AND user_id = ? AND analysis_id IS NULL",
            (reserved_log_id, user_id),
        )


def upload_eligibility(user: sqlite3.Row) -> Dict[str, Any]:
    """
    Indica si el usuario puede subir un nuevo análisis y qué mostrar si no.
    Devuelve: can_upload, limit_reason, invite_upgrade, invite_contact, contact_email.
    """
    uid = user["id"]
    plan = get_effective_plan(user)
    contact_email = next(iter(ADMIN_EMAILS), None) if ADMIN_EMAILS else None

    if plan == "free":
        month_gens = generations_this_month(uid)
        total_count = analyses_count(uid)
        if month_gens >= FREE_REPORTS_PER_MONTH:
            return {
                "can_upload": False,
                "limit_reason": "Ya usaste tu reporte gratuito de este mes.",
                "invite_upgrade": True,
                "invite_contact": False,
                "contact_email": contact_email,
            }
        if total_count >= FREE_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": f"Ya tienes {FREE_MAX_ANALYSES} análisis guardados (límite del plan gratis).",
                "invite_upgrade": True,
                "invite_contact": False,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    if plan == "pro":
        if generations_this_month(uid) >= PRO_90_REPORTS_PER_MONTH:
            return {
                "can_upload": False,
                "limit_reason": f"Has alcanzado el límite de lecturas de este mes en Pro ({PRO_90_REPORTS_PER_MONTH}).",
                "invite_upgrade": True,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        total_count = analyses_count(uid)
        if total_count >= PRO_90_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": "Has llegado al límite de análisis guardados de Pro.",
                "invite_upgrade": True,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    if plan == "pro_plus":
        if generations_this_month(uid) >= PRO_PLUS_REPORTS_PER_MONTH:
            return {
                "can_upload": False,
                "limit_reason": f"Has alcanzado el límite de lecturas de este mes en Pro+ ({PRO_PLUS_REPORTS_PER_MONTH}).",
                "invite_upgrade": False,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        total_count = analyses_count(uid)
        if total_count >= PRO_PLUS_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": "Has llegado al límite de análisis de tu plan.",
                "invite_upgrade": False,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    if plan == "free_trial":
        if generations_this_month(uid) >= PRO_PLUS_REPORTS_PER_MONTH:
            return {
                "can_upload": False,
                "limit_reason": f"Has alcanzado el límite de lecturas de este mes de tu prueba extendida ({PRO_PLUS_REPORTS_PER_MONTH}).",
                "invite_upgrade": False,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        total_count = analyses_count(uid)
        if total_count >= PRO_PLUS_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": "Has llegado al límite de análisis guardados de tu prueba extendida.",
                "invite_upgrade": True,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}


def _date_span_explanation(summary: Dict[str, Any], max_allowed: int, plan_label: str) -> str:
    od = int(summary.get("overall_days_covered") or 0)
    md = int(summary.get("max_days_covered") or 0)
    peak = max(od, md)
    head = (
        "No generamos la lectura: el rango de fechas de tus datos supera lo que incluye tu plan. "
        f"En {plan_label} el tope es de {max_allowed} días por análisis. "
    )
    absurd_threshold = max(max_allowed * 5, 730)
    if peak > absurd_threshold:
        return (
            head
            + "En esta carga el sistema calculó un rango de fechas muy amplio; suele deberse a una columna de fecha equivocada, "
            "formatos mezclados o valores atípicos en el export (no es usual en operación hotelera normal). "
            "El archivo que subiste no se modifica en el servidor. Revisa la columna de fechas en el PMS o en Excel, acota el export "
            "a las fechas que quieres leer o divide la carga en varios envíos; también puedes valorar ampliar plan si necesitas más días por lectura."
        )
    return (
        head
        + f"En esta carga detectamos una ventana de unos {od} día(s) en conjunto y hasta {md} día(s) en la serie más larga "
        f"(referencia: {peak} días). "
        "Tu export no se altera: detenemos la interpretación antes de consumir el análisis. "
        "Puedes acortar fechas en el PMS, enviar menos archivos en una sola corrida o subir de plan si aplica."
    )


def enforce_plan(user: sqlite3.Row, summary: Dict[str, Any]):
    uid = user["id"]
    plan = get_effective_plan(user)

    if plan == "free":
        if summary["total_files"] > FREE_MAX_FILES_PER_ANALYSIS:
            raise HTTPException(
                status_code=402,
                detail=(
                    "No generamos la lectura: en plan gratis va un solo archivo por corrida. "
                    "Quita los demás con ✕ o amplía plan; el sistema no modifica tus exports."
                ),
            )
        if summary["overall_days_covered"] > FREE_MAX_DAYS or summary["max_days_covered"] > FREE_MAX_DAYS:
            raise HTTPException(
                status_code=402,
                detail=_date_span_explanation(summary, FREE_MAX_DAYS, "plan gratis"),
            )
        if analyses_count(uid) >= FREE_MAX_ANALYSES:
            raise HTTPException(status_code=402, detail=f"Ya tienes {FREE_MAX_ANALYSES} análisis guardados (límite del plan gratis). Pásate a Pro para guardar más.")
        return

    if plan == "pro":
        if summary["total_files"] > PRO_90_MAX_FILES:
            raise HTTPException(
                status_code=402,
                detail=(
                    f"No generamos la lectura: el plan Pro admite hasta {PRO_90_MAX_FILES} archivos por corrida. "
                    "Reduce la carga o súbelo en otra sesión; tus archivos no se alteran."
                ),
            )
        if summary["overall_days_covered"] > PRO_90_MAX_DAYS or summary["max_days_covered"] > PRO_90_MAX_DAYS:
            raise HTTPException(
                status_code=402,
                detail=_date_span_explanation(summary, PRO_90_MAX_DAYS, "plan Pro")
                + " Pro+ permite hasta 180 días.",
            )
        if analyses_count(uid) >= PRO_90_MAX_ANALYSES:
            raise HTTPException(status_code=402, detail="Has llegado al límite de análisis guardados de Pro. Sube a Pro+ o contáctanos.")
        return

    if plan == "pro_plus":
        if summary["total_files"] > PRO_180_MAX_FILES:
            raise HTTPException(
                status_code=402,
                detail=(
                    f"No generamos la lectura: Pro+ admite hasta {PRO_180_MAX_FILES} archivos por corrida. "
                    "Parte la carga en dos envíos si necesitas más fuentes."
                ),
            )
        if summary["overall_days_covered"] > PRO_180_MAX_DAYS or summary["max_days_covered"] > PRO_180_MAX_DAYS:
            raise HTTPException(
                status_code=402,
                detail=_date_span_explanation(summary, PRO_180_MAX_DAYS, "Pro+")
                + " Contáctanos si necesitas ventanas más largas.",
            )
        if analyses_count(uid) >= PRO_PLUS_MAX_ANALYSES:
            raise HTTPException(status_code=402, detail="Has llegado al límite de análisis de tu plan. Contáctanos para ampliar tu capacidad.")
        return

    if plan == "free_trial":
        if summary["total_files"] > PRO_180_MAX_FILES:
            raise HTTPException(
                status_code=402,
                detail=(
                    f"No generamos la lectura: tu prueba extendida admite hasta {PRO_180_MAX_FILES} archivos por corrida "
                    "(mismo tope que Pro+). Parte la carga en dos envíos si necesitas más fuentes."
                ),
            )
        if analyses_count(uid) >= PRO_PLUS_MAX_ANALYSES:
            raise HTTPException(
                status_code=402,
                detail="Has llegado al límite de análisis guardados de tu prueba extendida. Contrata un plan o contáctanos.",
            )
        return


def save_analysis(
    user_id: int,
    title: str,
    plan: str,
    summary: Dict[str, Any],
    analysis: Dict[str, Any],
    files: List[UploadFile],
    reserved_run_log_id: Optional[int] = None,
) -> Tuple[int, str]:
    created_at = now_iso()
    share_token = secrets.token_urlsafe(24)
    with db() as conn:
        cur = conn.execute(
            """
            INSERT INTO analyses (user_id, title, plan_at_analysis, file_count, days_covered, summary_json, analysis_json, created_at, share_token)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                title,
                plan,
                len(files),
                int(summary.get("overall_days_covered", 0)),
                json.dumps(summary, ensure_ascii=False),
                json.dumps(analysis, ensure_ascii=False),
                created_at,
                share_token,
            ),
        )
        analysis_id = cur.lastrowid
        for upload in files:
            upload.file.seek(0)
            content = upload.file.read()
            if len(content) > MAX_UPLOAD_BYTES_PER_FILE:
                raise ValueError("Archivo excede el tamaño máximo permitido.")
            ext = (Path(upload.filename or "archivo").suffix or "").lower()
            if ext not in ALLOWED_UPLOAD_EXTENSIONS:
                ext = ".bin"
            safe_name = f"u{user_id}_a{analysis_id}_{secrets.token_hex(8)}{ext}"
            file_path = UPLOAD_DIR / safe_name
            file_path.write_bytes(content)
            conn.execute(
                "INSERT INTO uploaded_files (analysis_id, original_name, stored_path, created_at) VALUES (?, ?, ?, ?)",
                (analysis_id, upload.filename or safe_name, str(file_path), created_at),
            )
        if reserved_run_log_id:
            cur = conn.execute(
                """
                UPDATE analysis_run_log
                SET analysis_id = ?, created_at = ?
                WHERE id = ? AND user_id = ? AND analysis_id IS NULL
                """,
                (analysis_id, created_at, reserved_run_log_id, user_id),
            )
            if cur.rowcount != 1:
                raise ValueError("No se pudo enlazar la reserva de cupo mensual con el análisis guardado.")
        else:
            conn.execute(
                "INSERT INTO analysis_run_log (user_id, analysis_id, created_at) VALUES (?, ?, ?)",
                (user_id, analysis_id, created_at),
            )
        return analysis_id, share_token



def user_row_as_dict(row: sqlite3.Row) -> dict:
    """Convierte sqlite3.Row a dict para usar .get() y evitar AttributeError."""
    return dict(row) if row is not None else {}
