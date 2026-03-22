import hashlib
import hmac
import io
import json
import math
import os
import re
import secrets
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from threading import Lock
from typing import Any, Dict, List, Optional, Tuple
import time

from config import (
    ADMIN_EMAILS,
    ALLOWED_UPLOAD_EXTENSIONS,
    ANNUAL_PRICE,
    APP_NAME,
    APP_URL,
    BASE_DIR,
    DEFAULT_MODEL,
    FREE_MAX_ANALYSES,
    FREE_MAX_DAYS,
    FREE_MAX_FILES_PER_ANALYSIS,
    FREE_REPORTS_PER_MONTH,
    MAX_UPLOAD_BYTES_PER_FILE,
    MONTHLY_PRICE,
    OPENAI_API_KEY,
    PREMIUM_MONTHLY_PRICE,
    PRO_180_MAX_ANALYSES,
    PRO_180_MAX_DAYS,
    PRO_180_MAX_FILES,
    PRO_90_MAX_ANALYSES,
    PRO_90_MAX_DAYS,
    PRO_90_MAX_FILES,
    PRO_PLUS_MAX_ANALYSES,
    SECRET_KEY,
    SMTP_HOST,
    SMTP_PASSWORD,
    SMTP_USER,
    STRIPE_ANNUAL_PRICE_ID,
    STRIPE_MONTHLY_PRICE_ID,
    STRIPE_PRO_PLUS_PRICE_ID,
    STRIPE_PUBLISHABLE_KEY,
    STRIPE_SECRET_KEY,
    STRIPE_WEBHOOK_SECRET,
    TRIAL_DAYS,
    UPLOAD_DIR,
)
from db import db, init_db

# Esquema SQLite al importar el módulo (comportamiento previo a Fase 1).
init_db()

from debuglog import _dbg, _debug_log

"""
DragonApp — ensamblador FastAPI + lógica de negocio (análisis, billing, API v1).

Routers: routes/marketing, auth, consulting (aislado), admin, analysis. Ver docs/dragonapp_phase2.md.
"""

import pandas as pd
import requests
from auth_session import (
    get_api_user,
    is_admin_user,
    onboarding_pending,
    require_user,
)
from email_smtp import send_analysis_share_link_email
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, Response, UploadFile
from fastapi.exception_handlers import http_exception_handler as default_http_exception_handler
from fastapi.routing import APIRouter
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from plans import max_upload_files_for_plan, plan_label
from starlette.middleware.sessions import SessionMiddleware
from reportlab.lib.pagesizes import A4
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from templating import templates
from time_utils import now_iso


def public_share_base_url() -> str:
    return (APP_URL or "http://127.0.0.1:8000").rstrip("/")


app = FastAPI(
    title=APP_NAME,
    description="API para ejecutar análisis de reportes hoteleros, listar análisis y descargar PDFs. Autenticación por API key (header X-API-Key o Authorization: Bearer <key>).",
    version="1.0",
    docs_url="/docs",
    redoc_url=None,  # servimos ReDoc a mano con URL absoluta del schema para que cargue bien
    openapi_url="/openapi.json",
)
# Cookie de sesión: Secure en producción (HTTPS)
_https_only = (APP_URL or "").strip().lower().startswith("https")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, same_site="lax", https_only=_https_only)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Headers de seguridad para reducir riesgos XSS, clickjacking y fuga de referrer."""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Para 401 (no autenticado): redirigir a login en peticiones HTML; devolver JSON con redirect en API/fetch."""
    if exc.status_code == 401:
        accept = (request.headers.get("accept") or "").lower()
        wants_html = "text/html" in accept and "application/json" not in accept
        path = (request.url.path or "").strip()
        if wants_html or path.startswith("/app") or path.startswith("/admin") or path == "/":
            next_url = path if path and path != "/" else "/app"
            return RedirectResponse(url=f"/login?next={next_url}", status_code=303)
        return JSONResponse(
            {"ok": False, "error": exc.detail or "Debes iniciar sesión", "redirect": "/login"},
            status_code=401,
        )
    return await default_http_exception_handler(request, exc)


api_v1 = APIRouter(prefix="/api/v1", tags=["API v1"])

_SHARE_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def looks_like_email(addr: str) -> bool:
    s = (addr or "").strip()
    return bool(s) and len(s) <= 254 and bool(_SHARE_EMAIL_RE.match(s))


DATE_ALIASES = [
    "stay_date", "date", "business_date", "checkin", "check_in", "arrival", "arrival_date",
    "checkout", "check_out", "departure", "departure_date", "booking_date", "created_at",
    "reservation_date", "travel_date", "occupancy_date", "fecha", "fecha_noche", "dia",
]
REVENUE_ALIASES = [
    "net_room_revenue", "room_revenue", "revenue", "revenue_net", "net_revenue", "total_revenue",
    "room_rev", "amount", "booking_value", "room_amount", "total_amount",
    "ingresos", "ingreso", "ingreso_total", "revenue_total",
]
GROSS_ALIASES = ["gross_revenue", "gross_amount", "gross_booking_value", "sell_amount"]
COMM_ALIASES = ["commission", "commission_amount", "ota_commission", "channel_commission"]
PAYMENT_ALIASES = ["payment_cost", "card_fee", "payment_fee", "processing_fee"]
ROOM_NIGHTS_ALIASES = ["room_nights", "nights", "night_count", "roomnights"]
RESERVATION_ID_ALIASES = ["reservation_id", "booking_id", "confirmation", "conf_no", "res_id"]
CHANNEL_ALIASES = ["channel", "source", "booking_channel", "channel_name", "ota", "segmento_canal", "canal"]
STATUS_ALIASES = ["status", "reservation_status", "booking_status", "estatus"]
ROOMS_ALIASES = ["rooms", "room_count", "habitaciones"]
ADR_ALIASES = ["adr", "average_daily_rate", "tarifa_promedio"]
TAX_ALIASES = ["tax", "taxes", "vat", "impuestos"]


ANALYSIS_SCHEMA = {
    "type": "object",
    "properties": {
        "resumen_ejecutivo": {"type": "string"},
        "metricas_clave": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "nombre": {"type": "string"},
                    "valor": {"type": "string"},
                    "lectura": {"type": "string"},
                },
                "required": ["nombre", "valor", "lectura"],
                "additionalProperties": False,
            },
        },
        "hallazgos_prioritarios": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "titulo": {"type": "string"},
                    "detalle": {"type": "string"},
                    "impacto": {"type": "string"},
                    "prioridad": {"type": "string"},
                },
                "required": ["titulo", "detalle", "impacto", "prioridad"],
                "additionalProperties": False,
            },
        },
        "oportunidades_directo_vs_ota": {
            "type": "array",
            "items": {"type": "string"},
        },
        "riesgos_detectados": {
            "type": "array",
            "items": {"type": "string"},
        },
        "recomendaciones_accionables": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "accion": {"type": "string"},
                    "por_que": {"type": "string"},
                    "urgencia": {"type": "string"},
                },
                "required": ["accion", "por_que", "urgencia"],
                "additionalProperties": False,
            },
        },
        "datos_faltantes": {
            "type": "array",
            "items": {"type": "string"},
        },
        "senal_de_upgrade": {
            "type": "object",
            "properties": {
                "deberia_hacer_upgrade": {"type": "boolean"},
                "motivo": {"type": "string"},
            },
            "required": ["deberia_hacer_upgrade", "motivo"],
            "additionalProperties": False,
        },
    },
    "required": [
        "resumen_ejecutivo",
        "metricas_clave",
        "hallazgos_prioritarios",
        "oportunidades_directo_vs_ota",
        "riesgos_detectados",
        "recomendaciones_accionables",
        "datos_faltantes",
        "senal_de_upgrade",
    ],
    "additionalProperties": False,
}


def normalize_col(col: Any) -> str:
    col = str(col).strip().lower()
    col = re.sub(r"[^a-z0-9]+", "_", col)
    return col.strip("_")


def maybe_to_datetime(series: pd.Series) -> Optional[pd.Series]:
    try:
        converted = pd.to_datetime(series, errors="coerce")
        success_ratio = converted.notna().mean() if len(converted) else 0
        if success_ratio >= 0.45:
            return converted
    except Exception:
        return None
    return None


def find_col(columns: List[str], aliases: List[str]) -> Optional[str]:
    for alias in aliases:
        if alias in columns:
            return alias
    for col in columns:
        for alias in aliases:
            if alias in col:
                return col
    return None


def _has_date_header(columns: Any) -> bool:
    """True si las columnas parecen encabezado de datos (tienen Fecha/Día/Date)."""
    lower = [str(c).lower() for c in columns]
    return any("fecha" in c or "date" in c or "día" in c or "dia" in c for c in lower)


def _excel_to_dataframes_openpyxl(raw: bytes, filename: str, read_only: bool = True) -> List[Tuple[str, pd.DataFrame]]:
    """Fallback: leer Excel con openpyxl directo (soporta formatos que pandas ExcelFile a veces rechaza)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=read_only, data_only=True)
    result = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c) if c is not None else "" for c in rows[0]]
                data = rows[1:]
                df = pd.DataFrame(data, columns=header)
                df = df.dropna(how="all").reset_index(drop=True)
            else:
                df = pd.DataFrame()
            result.append((f"{filename}::{sheet_name}", df))
    finally:
        if read_only:
            wb.close()
    return result


def parse_file(upload: UploadFile) -> List[Tuple[str, pd.DataFrame]]:
    ext = os.path.splitext(upload.filename or "")[1].lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise ValueError(f"Formato no permitido. Solo: {', '.join(sorted(ALLOWED_UPLOAD_EXTENSIONS))}")
    raw = upload.file.read()
    if len(raw) > MAX_UPLOAD_BYTES_PER_FILE:
        raise ValueError(f"Archivo demasiado grande. Máximo {MAX_UPLOAD_BYTES_PER_FILE // (1024*1024)} MB por archivo.")
    if ext == ".csv":
        # Detectar separador: coma o punto y coma (común en Excel en español)
        buf = io.BytesIO(raw)
        first_line = buf.readline().decode("utf-8", errors="replace").strip()
        buf.seek(0)
        sep = ","
        if ";" in first_line and "," not in first_line:
            sep = ";"
        elif ";" in first_line and first_line.count(";") >= first_line.count(","):
            sep = ";"
        try:
            df = pd.read_csv(buf, sep=sep, encoding="utf-8")
        except Exception:
            df = pd.read_csv(io.BytesIO(raw), encoding="utf-8")
        # Si la primera fila parece encabezado de PMS (título, moneda, etc.), buscar fila con Fecha/Día
        if not _has_date_header(df.columns) and len(df) > 0:
            for skip in range(1, min(8, len(raw) // 80 + 1)):
                try:
                    buf2 = io.BytesIO(raw)
                    df2 = pd.read_csv(buf2, sep=sep, encoding="utf-8", skiprows=skip)
                    if len(df2.columns) >= 2 and _has_date_header(df2.columns):
                        df = df2
                        break
                except Exception:
                    continue
        return [(upload.filename or "reporte.csv", df)]
    if ext == ".xls":
        # Excel 97-2003: xlrd a veces marca como "corruptos" archivos válidos de PMS; ignorar esa comprobación
        try:
            excel = pd.ExcelFile(
                io.BytesIO(raw),
                engine="xlrd",
                engine_kwargs={"ignore_workbook_corruption": True},
            )
            result = []
            for sheet in excel.sheet_names:
                df = excel.parse(sheet)
                if not _has_date_header(df.columns) and len(df) >= 4:
                    for header_row in range(1, 6):
                        try:
                            df2 = excel.parse(sheet, header=header_row)
                            if len(df2.columns) >= 2 and _has_date_header(df2.columns):
                                df = df2
                                break
                        except Exception:
                            continue
                result.append((f"{upload.filename}::{sheet}", df))
            return result
        except Exception as e:
            raise ValueError(
                "No pudimos leer este archivo Excel 97-2003 (.xls). "
                "Prueba exportando el reporte en CSV o en formato .xlsx."
            ) from e
    if ext in [".xlsx", ".xlsm"]:
        # Intentar primero con pandas (rápido y estándar)
        try:
            excel = pd.ExcelFile(io.BytesIO(raw))
            result = []
            for sheet in excel.sheet_names:
                df = excel.parse(sheet)
                if not _has_date_header(df.columns) and len(df) >= 4:
                    for header_row in range(1, 6):
                        try:
                            df2 = excel.parse(sheet, header=header_row)
                            if len(df2.columns) >= 2 and _has_date_header(df2.columns):
                                df = df2
                                break
                        except Exception:
                            continue
                result.append((f"{upload.filename}::{sheet}", df))
            return result
        except Exception as e:
            err = str(e).strip().lower()
            # Si falla por “corruption” o formato raro, intentar con openpyxl en modo read_only (otra ruta de lectura)
            if "corruption" in err or "workbook" in err or "seen[" in err or "bad zip" in err or "invalid" in err:
                for use_read_only in (True, False):
                    try:
                        return _excel_to_dataframes_openpyxl(raw, upload.filename or "reporte.xlsx", read_only=use_read_only)
                    except Exception:
                        continue
                raise ValueError(
                    "No pudimos leer este Excel (formato propio de tu PMS/channel). "
                    "Estamos mejorando el soporte. Mientras tanto: exporta el mismo reporte en CSV y súbelo aquí."
                ) from e
            raise
    raise ValueError(f"Formato no soportado: {ext}")


def infer_sheet(sheet_name: str, df: pd.DataFrame) -> Dict[str, Any]:
    original_cols = [str(c) for c in df.columns]
    norm_cols = [normalize_col(c) for c in original_cols]
    work = df.copy()
    work.columns = norm_cols
    work = work.dropna(how="all")
    if len(work) > 5000:
        work = work.head(5000)

    mappings = {
        "reservation_id": find_col(norm_cols, RESERVATION_ID_ALIASES),
        "channel": find_col(norm_cols, CHANNEL_ALIASES),
        "status": find_col(norm_cols, STATUS_ALIASES),
        "revenue": find_col(norm_cols, REVENUE_ALIASES),
        "gross_revenue": find_col(norm_cols, GROSS_ALIASES),
        "commission": find_col(norm_cols, COMM_ALIASES),
        "payment_cost": find_col(norm_cols, PAYMENT_ALIASES),
        "room_nights": find_col(norm_cols, ROOM_NIGHTS_ALIASES),
        "rooms": find_col(norm_cols, ROOMS_ALIASES),
        "adr": find_col(norm_cols, ADR_ALIASES),
        "taxes": find_col(norm_cols, TAX_ALIASES),
    }

    date_candidates = []
    for col in norm_cols:
        if any(alias in col for alias in DATE_ALIASES):
            converted = maybe_to_datetime(work[col])
            if converted is not None:
                work[col] = converted
                date_candidates.append(col)

    for col in norm_cols:
        if col not in date_candidates and work[col].dtype == object:
            converted = maybe_to_datetime(work[col])
            if converted is not None and len(date_candidates) < 4:
                work[col] = converted
                date_candidates.append(col)

    checkin_col = next((c for c in date_candidates if any(x in c for x in ["checkin", "check_in", "arrival"])), None)
    checkout_col = next((c for c in date_candidates if any(x in c for x in ["checkout", "check_out", "departure"])), None)
    stay_col = next((c for c in date_candidates if any(x in c for x in ["stay", "business_date", "occupancy"])), None) or checkin_col
    booking_col = next((c for c in date_candidates if any(x in c for x in ["booking", "created", "reservation_date"])), None)

    for col in norm_cols:
        try:
            converted = pd.to_numeric(work[col], errors="coerce")
            if converted.notna().mean() >= 0.5:
                work[col] = converted
        except Exception:
            pass

    row_count = int(len(work))
    rev_col = mappings["revenue"] or mappings["gross_revenue"]
    comm_col = mappings["commission"]
    pay_col = mappings["payment_cost"]
    rn_col = mappings["room_nights"]
    channel_col = mappings["channel"]
    status_col = mappings["status"]
    adr_col = mappings["adr"]

    # Convertir columnas de dinero con formato "$1,234.56" a numérico
    def _to_numeric_currency(series: pd.Series) -> pd.Series:
        if series.dtype != object and pd.api.types.is_numeric_dtype(series):
            return series
        cleaned = series.astype(str).str.replace(r"[$,\s]", "", regex=True)
        return pd.to_numeric(cleaned, errors="coerce")

    for col in [rev_col, comm_col, pay_col, adr_col]:
        if col and col in work.columns:
            work[col] = _to_numeric_currency(work[col])

    metrics: Dict[str, Any] = {}
    if rev_col:
        metrics["revenue_total"] = round(float(work[rev_col].fillna(0).sum()), 2)
    if comm_col:
        metrics["comision_total"] = round(float(work[comm_col].fillna(0).sum()), 2)
    if pay_col:
        metrics["costo_pago_total"] = round(float(work[pay_col].fillna(0).sum()), 2)
    if rn_col:
        metrics["room_nights"] = round(float(work[rn_col].fillna(0).sum()), 2)
    elif checkin_col and checkout_col:
        metrics["room_nights"] = round(float((work[checkout_col] - work[checkin_col]).dt.days.clip(lower=0).fillna(0).sum()), 2)
    if adr_col:
        metrics["adr_promedio"] = round(float(pd.to_numeric(work[adr_col], errors="coerce").mean()), 2)
    elif metrics.get("revenue_total") and metrics.get("room_nights"):
        if metrics["room_nights"] > 0:
            metrics["adr_estimado"] = round(metrics["revenue_total"] / metrics["room_nights"], 2)

    if status_col:
        s = work[status_col].astype(str).str.lower()
        cancelled = int(s.str.contains("cancel|void|no show|noshow").sum())
        metrics["cancelaciones"] = cancelled
        metrics["cancelacion_pct"] = round((cancelled / max(row_count, 1)) * 100, 2)

    if channel_col:
        channel_series = work[channel_col].astype(str).str.strip().replace("", "Sin canal")
        top_channels = channel_series.value_counts().head(8)
        metrics["top_canales_por_reservas"] = [{"canal": str(k), "reservas": int(v)} for k, v in top_channels.items()]
        if rev_col:
            rev_by_channel = work.groupby(channel_series)[rev_col].sum().sort_values(ascending=False).head(8)
            metrics["top_canales_por_ingreso"] = [{"canal": str(k), "ingreso": round(float(v), 2)} for k, v in rev_by_channel.items()]
        if comm_col and rev_col:
            grouped = work.groupby(channel_series)[[rev_col, comm_col]].sum()
            channel_margin = []
            for chan, row in grouped.iterrows():
                margin = float(row[rev_col]) - float(row[comm_col])
                channel_margin.append({"canal": str(chan), "ingreso": round(float(row[rev_col]), 2), "comision": round(float(row[comm_col]), 2), "margen_estimado": round(margin, 2)})
            metrics["margen_estimado_por_canal"] = sorted(channel_margin, key=lambda x: x["margen_estimado"], reverse=True)[:8]

    min_date = None
    max_date = None
    for col in [stay_col, checkin_col, booking_col, checkout_col]:
        if col and col in work.columns:
            col_min = work[col].dropna().min()
            col_max = work[col].dropna().max()
            if col_min is not pd.NaT and col_max is not pd.NaT:
                if min_date is None or col_min < min_date:
                    min_date = col_min
                if max_date is None or col_max > max_date:
                    max_date = col_max
    # Reportes tipo forecast/revenue con una columna "Fecha" (una fila por día): usar primera columna de fecha
    if (min_date is None or max_date is None) and date_candidates:
        for col in date_candidates:
            if col not in work.columns:
                continue
            col_min = work[col].dropna().min()
            col_max = work[col].dropna().max()
            if col_min is not pd.NaT and col_max is not pd.NaT:
                if min_date is None or col_min < min_date:
                    min_date = col_min
                if max_date is None or col_max > max_date:
                    max_date = col_max
                break

    # Normalizar a tipo fecha (evitar numpy.int64 p. ej. fechas Excel como número)
    try:
        min_date_norm = pd.Timestamp(min_date) if min_date is not None else None
    except (TypeError, ValueError):
        min_date_norm = None
    try:
        max_date_norm = pd.Timestamp(max_date) if max_date is not None else None
    except (TypeError, ValueError):
        max_date_norm = None

    days_covered = 0
    if min_date_norm is not None and max_date_norm is not None:
        try:
            delta = max_date_norm - min_date_norm
            days_covered = int(getattr(delta, "days", delta)) + 1
        except (TypeError, ValueError):
            days_covered = 0

    fields_detected = [k for k, v in mappings.items() if v]
    return {
        "sheet_name": sheet_name,
        "rows": row_count,
        "mappings": mappings,
        "date_columns": {
            "stay": stay_col,
            "booking": booking_col,
            "checkin": checkin_col,
            "checkout": checkout_col,
        },
        "fields_detected": fields_detected,
        "metrics": metrics,
        "days_covered": days_covered,
        "date_range": {
            "start": min_date_norm.isoformat() if min_date_norm is not None else None,
            "end": max_date_norm.isoformat() if max_date_norm is not None else None,
        },
        "sample_columns": norm_cols[:30],
    }


def summarize_reports(files: List[UploadFile]) -> Dict[str, Any]:
    report_summaries = []
    max_days = 0
    all_starts = []
    all_ends = []
    for upload in files:
        sheets = parse_file(upload)
        for sheet_name, df in sheets:
            summary = infer_sheet(sheet_name, df)
            report_summaries.append(summary)
            max_days = max(max_days, summary.get("days_covered", 0))
            dr = summary.get("date_range", {})
            if dr.get("start"):
                t = pd.to_datetime(dr["start"])
                if pd.notna(t):
                    all_starts.append(t)
            if dr.get("end"):
                t = pd.to_datetime(dr["end"])
                if pd.notna(t):
                    all_ends.append(t)
    overall_days = 0
    if all_starts and all_ends:
        try:
            delta = max(all_ends) - min(all_starts)
            d = getattr(delta, "days", None)
            if d is not None and not (isinstance(d, float) and math.isnan(d)):
                overall_days = int(d) + 1
        except (TypeError, ValueError):
            pass
    return {
        "total_files": len(files),
        "reports_detected": len(report_summaries),
        "max_days_covered": max_days,
        "overall_days_covered": overall_days,
        "report_summaries": report_summaries,
    }


HOTEL_PROMPT = """
Eres DRAGONNÉ, director de revenue, distribución y e‑commerce hotelero con experiencia en cadenas internacionales. Siempre respondes en español LATAM, con lenguaje hotelero real y tono de consultor senior (claro, directo, accionable, cero “data por decir datos”).

Recibirás un JSON con:
- plan: "free_30", "pro_90" o "pro_180".
- contexto_hotel: objeto con hotel_nombre, hotel_tamano, hotel_categoria, hotel_ubicacion, hotel_estrellas, hotel_ubicacion_destino; además (opcionales): hotel_pms, hotel_channel_manager, hotel_booking_engine, hotel_tech_other (stack tecnológico), hotel_google_business_url, hotel_expedia_url, hotel_booking_url (enlaces de presencia online). Úsalo SIEMPRE para personalizar.
- contexto_negocio: qué le preocupa al usuario.
- resumen: lectura estructurada de reportes exportados desde PMS / channel / motor (CSV/Excel), incluyendo métricas por canal, días cubiertos y fechas.

Objetivo: entregar una lectura que un dueño / GM entienda en 5 minutos y que un revenue manager pueda usar para decidir hoy qué mover en precio, canales y estrategia.

Reglas de estilo general:
1) No repitas el reporte: tu trabajo NO es decir “hay X reservas” sino explicar QUÉ SIGNIFICA para el negocio. Cada vez que menciones un número, acompáñalo de una lectura (“esto indica…”, “esto implica riesgo de…”).
2) Escribe como colega senior: directo, honesto y empático. Si ves un problema serio (dependencia extrema de una OTA, precios muy bajos en fines de semana fuertes, ocupación de lunes–jueves muy floja), dilo con claridad y con sentido de urgencia.
3) Estructura siempre la respuesta en bloques claros del JSON: resumen_ejecutivo, hallazgos_prioritarios, recomendaciones_accionables, datos_faltantes. Dentro de cada bloque, prioriza 3–7 puntos potentes, no listas largas de detalles sin priorizar.

Reglas analíticas (cómo leer los datos):
4) Palancas que deben aparecer SIEMPRE que haya datos suficientes:
   - Mezcla de canales: canal directo vs OTAs, dependencia de 1–2 agencias, canales con mucho volumen pero ADR o margen neto claramente más bajo.
   - Pricing y ADR: diferencias de tarifa entre canales para las mismas fechas, entre semana vs fin de semana, y entre tipos de día (p.ej. noches con eventos/local demand drivers si las detectas).
   - Ritmo y distribución en el tiempo: qué días de la semana cargan la ocupación (¿solo viernes-sábado?, ¿se cae domingo-jueves?), tendencias dentro del rango (primer tercio vs último tercio).
   - Coste de distribución: cuando haya comisiones o % estimados, habla de margen neto y no solo de revenue bruto; si no hay comisiones, dilo y explica por qué sería clave tenerlas.
   - Riesgos y oportunidades de inventario: noches con poca demanda donde tiene sentido empujar volumen, y noches fuertes donde conviene proteger tarifa y mix directo.
5) Cuando veas patrones interesantes, dilo explícitamente. Ejemplos de redacción:
   - “Tu ocupación se concentra casi solo en viernes y sábado; de domingo a jueves el hotel funciona a media máquina.”
   - “Booking.com te aporta la mayor producción, pero a un ADR un 12–15% más bajo que el resto de tus canales en las mismas fechas; eso está erosionando margen.”
   - “El canal directo casi no participa en fines de semana fuertes; estás dejando toda la fecha en manos de OTAs con comisión.”

Uso del contexto del hotel:
6) CONTEXTO DEL HOTEL (obligatorio): El resumen_ejecutivo y las recomendaciones_accionables deben estar explícitamente adaptados al tipo de propiedad (hotel_tamano), categoría (hotel_categoria), estrellas (hotel_estrellas si > 0), ubicación geográfica (hotel_ubicacion) y posición en el destino (hotel_ubicacion_destino: centro, zona turística o periferia). No des el mismo consejo a un boutique que a un resort, ni a un hotel de ciudad que a un all-inclusive; adapta tono y expectativas a estrellas y ubicación (centro vs periferia).
7) Usa el stack tecnológico (PMS, channel manager, booking engine, otras tecnologías) para proponer acciones realistas: por ejemplo, hablar de reglas en el channel manager, ajustes de inventario hacia canal directo, paquetes en el motor de reservas, etc., siempre que tenga sentido.
8) IMPORTANTE sobre enlaces (Google My Business, Expedia, Booking): los URLs en contexto_hotel son SOLO para contexto (saber dónde está el hotel online). NUNCA inventes datos, reseñas ni métricas a partir de esas URLs; usa solo lo que viene en el resumen de reportes subidos.

Recomendaciones y enfoque accionable:
9) Recomendaciones: nada genérico tipo "revisar pricing". Sé específico en qué moverías (tarifa, canal, estancia mínima, política de cancelación, bundles, visibilidad por canal), en qué fechas / rangos de días / tipos de día y con qué objetivo (subir margen, ganar directo, reducir dependencia de OTA, etc.); y que sean coherentes con la categoría y tipo de propiedad del hotel.
10) Siempre que sea posible, traduce el hallazgo en un “plan de juego” para las próximas 2–4 semanas, agrupando acciones por prioridad: qué moverías primero, qué dejarías en observación y qué revisarías cuando haya más datos.

Datos faltantes y uso responsable:
11) No inventes datos; si asumes algo, dilo. Marca métricas estimadas como tales y mantenlas conservadoras.
12) Datos faltantes: di con claridad qué falta (ej. canales, comisiones, room nights, fechas de estancia, segmentos) y qué tipo de reporte adicional debería subir el usuario. No uses la falta de datos como excusa para no extraer valor de lo que sí existe.
13) REPORTES SIN CANALES (forecast, revenue por día): Si el resumen no incluye canales de distribución pero SÍ tiene ingresos, ADR y/o fechas (p. ej. reporte Forecast and Revenue con una fila por día), OBLIGATORIO extraer valor de lo que sí hay. Analiza: ingresos y ADR por día de la semana, diferencia entre semana vs fin de semana, tendencia en el mes, oportunidades de precio entre semana. Da recomendaciones concretas usando esos números. En resumen_ejecutivo incluye una línea tipo: "Esto es lo que podemos rescatar de tu reporte; con reportes que incluyan canales de distribución podemos aportar aún más valor (mix directo vs OTA, comisiones)." En datos_faltantes menciona canales/comisiones como mejora futura, no como excusa para no analizar lo que sí está.

Planes y señal de upgrade:
14) Planes:
   - "free_30": trata el caso como diagnóstico rápido sobre un máximo de 30 días; prioriza quick wins y, si faltan periodos para tener contexto, explícalo en la señal de upgrade.
   - "pro_90" y "pro_180": además de la foto actual, compara dentro del rango primeros días vs últimos (por ejemplo, primer tercio vs último tercio) y marca en hallazgos_prioritarios cualquier cambio de tendencia relevante (mix de canal, ADR, margen, cancelaciones). Usa también recomendaciones_accionables para proponer cómo reaccionar a esos cambios.
15) Señal de upgrade: si con más días o más reportes la lectura sería mucho más fuerte, dilo con tacto y con un ejemplo concreto (ej. "comparar este trimestre con el mismo periodo del año pasado").
16) Las integraciones directas con PMS (Mews, Cloudbeds, Little Hotelier) están pensadas como "Próximamente"; no prometas nada distinto a eso.

Devuelve SIEMPRE exclusivamente el JSON con el esquema especificado por el backend, sin texto adicional.
""".strip()


def call_openai(summary: Dict[str, Any], business_context: str, hotel_context: Dict[str, Any], plan: str) -> Dict[str, Any]:
    if not OPENAI_API_KEY:
        raise RuntimeError("No encontré OPENAI_API_KEY en el backend.")

    payload = {
        "model": DEFAULT_MODEL,
        "input": [
            {
                "role": "system",
                "content": [{"type": "input_text", "text": HOTEL_PROMPT}],
            },
            {
                "role": "user",
                "content": [{
                    "type": "input_text",
                    "text": json.dumps({
                        "plan": plan,
                        "contexto_hotel": hotel_context,
                        "uso_contexto_hotel": "Usa siempre contexto_hotel (nombre, tipo de propiedad/tamaño, categoría, ubicación) para redactar resumen_ejecutivo y recomendaciones_accionables de forma específica a este hotel.",
                        "contexto_negocio": business_context or "No se proporcionó contexto adicional.",
                        "resumen": summary,
                    }, ensure_ascii=False),
                }],
            },
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "profitpilot_hotel_analysis",
                "schema": ANALYSIS_SCHEMA,
                "strict": True,
            }
        },
    }
    response = requests.post(
        "https://api.openai.com/v1/responses",
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=120,
    )
    response.raise_for_status()
    data = response.json()
    text = None
    if data.get("output"):
        for item in data["output"]:
            for content in item.get("content", []):
                if content.get("type") in {"output_text", "text"} and content.get("text"):
                    text = content["text"]
                    break
            if text:
                break
    if not text:
        text = data.get("output_text") or data.get("text", "")
    if not text:
        raise RuntimeError("No pude leer la respuesta estructurada de OpenAI.")
    return json.loads(text)


def analyses_count(user_id: int) -> int:
    with db() as conn:
        row = conn.execute("SELECT COUNT(*) AS c FROM analyses WHERE user_id = ?", (user_id,)).fetchone()
        return int(row["c"])


def analyses_this_month(user_id: int) -> int:
    """Análisis creados en el mes actual (UTC) para el usuario."""
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()[:19]
    with db() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM analyses WHERE user_id = ? AND created_at >= ?",
            (user_id, start_of_month),
        ).fetchone()
        return int(row["c"])


def upload_eligibility(user: sqlite3.Row) -> Dict[str, Any]:
    """
    Indica si el usuario puede subir un nuevo análisis y qué mostrar si no.
    Devuelve: can_upload, limit_reason, invite_upgrade, invite_contact, contact_email.
    """
    uid = user["id"]
    plan = user["plan"]
    contact_email = next(iter(ADMIN_EMAILS), None) if ADMIN_EMAILS else None

    if plan == "free":
        month_count = analyses_this_month(uid)
        total_count = analyses_count(uid)
        if month_count >= FREE_REPORTS_PER_MONTH:
            return {
                "can_upload": False,
                "limit_reason": "Ya usaste tu reporte gratuito de este mes.",
                "invite_upgrade": True,
                "invite_contact": False,
                "contact_email": contact_email,
            }
        if total_count >= FREE_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": f"Ya tienes {FREE_MAX_ANALYSES} análisis guardados (límite del plan gratis).",
                "invite_upgrade": True,
                "invite_contact": False,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    if plan == "pro":
        total_count = analyses_count(uid)
        if total_count >= PRO_90_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": "Has llegado al límite de análisis guardados de Pro.",
                "invite_upgrade": True,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    if plan == "pro_plus":
        total_count = analyses_count(uid)
        if total_count >= PRO_PLUS_MAX_ANALYSES:
            return {
                "can_upload": False,
                "limit_reason": "Has llegado al límite de análisis de tu plan.",
                "invite_upgrade": False,
                "invite_contact": True,
                "contact_email": contact_email,
            }
        return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}

    return {"can_upload": True, "limit_reason": None, "invite_upgrade": False, "invite_contact": False, "contact_email": contact_email}


def enforce_plan(user: sqlite3.Row, summary: Dict[str, Any]):
    uid = user["id"]
    plan = user["plan"]

    if plan == "free":
        if summary["total_files"] > FREE_MAX_FILES_PER_ANALYSIS:
            raise HTTPException(status_code=402, detail=f"El plan gratis permite hasta {FREE_MAX_FILES_PER_ANALYSIS} archivos por análisis.")
        if summary["overall_days_covered"] > FREE_MAX_DAYS or summary["max_days_covered"] > FREE_MAX_DAYS:
            raise HTTPException(status_code=402, detail=f"El plan gratis permite interpretar hasta {FREE_MAX_DAYS} días de histórico o futuro por análisis.")
        if analyses_this_month(uid) >= FREE_REPORTS_PER_MONTH:
            raise HTTPException(status_code=402, detail="Ya usaste tu reporte gratuito de este mes. Pásate a Pro para seguir subiendo.")
        if analyses_count(uid) >= FREE_MAX_ANALYSES:
            raise HTTPException(status_code=402, detail=f"Ya tienes {FREE_MAX_ANALYSES} análisis guardados (límite del plan gratis). Pásate a Pro para guardar más.")
        return

    if plan == "pro":
        if summary["total_files"] > PRO_90_MAX_FILES:
            raise HTTPException(status_code=402, detail=f"El plan Pro permite hasta {PRO_90_MAX_FILES} archivos por análisis.")
        if summary["overall_days_covered"] > PRO_90_MAX_DAYS or summary["max_days_covered"] > PRO_90_MAX_DAYS:
            raise HTTPException(status_code=402, detail=f"El plan Pro permite hasta {PRO_90_MAX_DAYS} días por análisis. Sube a Pro+ para 180 días.")
        if analyses_count(uid) >= PRO_90_MAX_ANALYSES:
            raise HTTPException(status_code=402, detail="Has llegado al límite de análisis guardados de Pro. Sube a Pro+ o contáctanos.")
        return

    if plan == "pro_plus":
        if summary["total_files"] > PRO_180_MAX_FILES:
            raise HTTPException(status_code=402, detail=f"El plan Pro+ permite hasta {PRO_180_MAX_FILES} archivos por análisis.")
        if summary["overall_days_covered"] > PRO_180_MAX_DAYS or summary["max_days_covered"] > PRO_180_MAX_DAYS:
            raise HTTPException(status_code=402, detail=f"El plan Pro+ permite hasta {PRO_180_MAX_DAYS} días por análisis.")
        if analyses_count(uid) >= PRO_PLUS_MAX_ANALYSES:
            raise HTTPException(status_code=402, detail="Has llegado al límite de análisis de tu plan. Contáctanos para ampliar tu capacidad.")
        return


def save_analysis(user_id: int, title: str, plan: str, summary: Dict[str, Any], analysis: Dict[str, Any], files: List[UploadFile]) -> Tuple[int, str]:
    created_at = now_iso()
    share_token = secrets.token_urlsafe(24)
    with db() as conn:
        cur = conn.execute(
            """
            INSERT INTO analyses (user_id, title, plan_at_analysis, file_count, days_covered, summary_json, analysis_json, created_at, share_token)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                title,
                plan,
                len(files),
                int(summary.get("overall_days_covered", 0)),
                json.dumps(summary, ensure_ascii=False),
                json.dumps(analysis, ensure_ascii=False),
                created_at,
                share_token,
            ),
        )
        analysis_id = cur.lastrowid
        for upload in files:
            upload.file.seek(0)
            content = upload.file.read()
            if len(content) > MAX_UPLOAD_BYTES_PER_FILE:
                raise ValueError("Archivo excede el tamaño máximo permitido.")
            ext = (Path(upload.filename or "archivo").suffix or "").lower()
            if ext not in ALLOWED_UPLOAD_EXTENSIONS:
                ext = ".bin"
            safe_name = f"u{user_id}_a{analysis_id}_{secrets.token_hex(8)}{ext}"
            file_path = UPLOAD_DIR / safe_name
            file_path.write_bytes(content)
            conn.execute(
                "INSERT INTO uploaded_files (analysis_id, original_name, stored_path, created_at) VALUES (?, ?, ?, ?)",
                (analysis_id, upload.filename or safe_name, str(file_path), created_at),
            )
        return analysis_id, share_token


def format_money(value: float) -> str:
    return f"${value:,.2f}"


def stripe_request(method: str, path: str, data: Dict[str, Any]) -> Dict[str, Any]:
    if not STRIPE_SECRET_KEY:
        raise RuntimeError("Falta STRIPE_SECRET_KEY")
    response = requests.request(
        method,
        f"https://api.stripe.com{path}",
        auth=(STRIPE_SECRET_KEY, ""),
        data=data,
        timeout=60,
    )
    response.raise_for_status()
    return response.json()


def ensure_stripe_customer(user: sqlite3.Row) -> str:
    if user["stripe_customer_id"]:
        return user["stripe_customer_id"]
    customer = stripe_request("POST", "/v1/customers", {
        "email": user["email"],
        "name": user["contact_name"],
        "metadata[user_id]": str(user["id"]),
        "metadata[hotel_name]": user["hotel_name"],
    })
    customer_id = customer["id"]
    with db() as conn:
        conn.execute("UPDATE users SET stripe_customer_id = ?, updated_at = ? WHERE id = ?", (customer_id, now_iso(), user["id"]))
    return customer_id


def sync_user_from_stripe_customer(customer_id: str, subscription_id: Optional[str], status: str):
    plan = "free"
    if status in {"active", "trialing", "past_due"} and subscription_id:
        try:
            sub = stripe_request("GET", f"/v1/subscriptions/{subscription_id}")
            items = sub.get("items", {}).get("data", [])
            if items:
                price_obj = items[0].get("price")
                price_id = price_obj.get("id", "") if isinstance(price_obj, dict) else (price_obj or "")
                if price_id == STRIPE_PRO_PLUS_PRICE_ID:
                    plan = "pro_plus"
                elif price_id in (STRIPE_MONTHLY_PRICE_ID, STRIPE_ANNUAL_PRICE_ID):
                    plan = "pro"
                else:
                    plan = "pro"
        except Exception:
            plan = "pro"
    elif status in {"active", "trialing", "past_due"}:
        plan = "pro"
    with db() as conn:
        conn.execute(
            "UPDATE users SET plan = ?, stripe_subscription_id = ?, updated_at = ? WHERE stripe_customer_id = ?",
            (plan, subscription_id, now_iso(), customer_id),
        )


@app.get("/app/account", response_class=HTMLResponse)
def account_page(request: Request):
    """Mi cuenta: perfil, editar datos, y opciones de plan (Pro y Pro+)."""
    user = require_user(request)
    if onboarding_pending(user):
        return RedirectResponse("/onboarding", status_code=303)
    return templates.TemplateResponse("account.html", {
        "request": request,
        "user": user,
        "plan_label": plan_label(user["plan"]),
        "monthly_price": MONTHLY_PRICE,
        "premium_monthly_price": PREMIUM_MONTHLY_PRICE,
        "stripe_publishable_key": STRIPE_PUBLISHABLE_KEY,
    })


@app.get("/app", response_class=HTMLResponse)
def dashboard(request: Request):
    user = require_user(request)
    if onboarding_pending(user):
        return RedirectResponse("/onboarding", status_code=303)
    # Tracking de sesión básica: último uso del panel
    session_id = request.session.get("session_id")
    if session_id:
        with db() as conn:
            conn.execute(
                "UPDATE user_sessions SET last_seen_at = ?, request_count = request_count + 1 WHERE id = ?",
                (now_iso(), session_id),
            )
    with db() as conn:
        analyses = conn.execute("SELECT * FROM analyses WHERE user_id = ? ORDER BY created_at DESC LIMIT 20", (user["id"],)).fetchall()
    formatted = []
    for row in analyses:
        # #region agent log
        _created = row["created_at"]
        _days = row["days_covered"]
        _debug_log("app.py:dashboard", "analysis row", {"id": row["id"], "created_at_is_none": _created is None, "days_covered_is_none": _days is None}, "H1")
        # #endregion
        analysis = json.loads(row["analysis_json"])
        summary = json.loads(row["summary_json"])
        created_at_str = (_created[:19].replace("T", " ") if _created else "")
        formatted.append({
            "id": row["id"],
            "title": row["title"] or f"Análisis {row['id']}",
            "created_at": created_at_str,
            "file_count": row["file_count"],
            "days_covered": _days if _days is not None else 0,
            "resumen_ejecutivo": analysis.get("resumen_ejecutivo", ""),
            "metricas": analysis.get("metricas_clave", [])[:4],
            "senal_upgrade": analysis.get("senal_de_upgrade", {}),
            "reports_detected": summary.get("reports_detected", 0),
        })
    eligibility = upload_eligibility(user)
    return templates.TemplateResponse("app.html", {
        "request": request,
        "user": user,
        "is_admin": is_admin_user(user),
        "analyses": formatted,
        "plan_label": plan_label(user["plan"]),
        "max_files_per_analysis": max_upload_files_for_plan(user["plan"]),
        "pro_max_files": PRO_90_MAX_FILES,
        "pro_plus_max_files": PRO_180_MAX_FILES,
        "free_max_days": FREE_MAX_DAYS,
        "free_max_files": FREE_MAX_FILES_PER_ANALYSIS,
        "free_max_analyses": FREE_MAX_ANALYSES,
        "monthly_price": MONTHLY_PRICE,
        "annual_price": ANNUAL_PRICE,
        "premium_monthly_price": PREMIUM_MONTHLY_PRICE,
        "stripe_publishable_key": STRIPE_PUBLISHABLE_KEY,
        "can_upload": eligibility["can_upload"],
        "limit_reason": eligibility["limit_reason"],
        "invite_upgrade": eligibility["invite_upgrade"],
        "invite_contact": eligibility["invite_contact"],
        "contact_email": eligibility["contact_email"],
        "smtp_configured": bool(SMTP_HOST and SMTP_USER and SMTP_PASSWORD),
    })


def _user_row_as_dict(row: sqlite3.Row) -> dict:
    """Convierte sqlite3.Row a dict para usar .get() y evitar AttributeError."""
    return dict(row) if row is not None else {}


async def web_analyze(request: Request, business_context: str, files: List[UploadFile]):
    # #region agent log
    _debug_log("app.py:analyze", "POST /analyze entry", {"files_count": len(files) if files else 0}, "H4")
    _dbg("app.py:analyze", "entry", {"files_count": len(files) if files else 0, "filenames": [getattr(f, "filename", None) for f in (files or [])]}, "H_A")
    # #endregion
    user = _user_row_as_dict(require_user(request))
    if onboarding_pending(user):
        _dbg("app.py:analyze", "early_return", {"reason": "onboarding_pending"}, "H_C")
        return JSONResponse({"ok": False, "error": "Completa los datos de tu hotel primero.", "redirect": "/onboarding"}, status_code=400)
    if not files:
        _dbg("app.py:analyze", "early_return", {"reason": "no_files"}, "H_A")
        return JSONResponse({"ok": False, "error": "Sube al menos un reporte."}, status_code=400)
    try:
        summary = summarize_reports(files)
        _dbg("app.py:analyze", "after_summarize", {"reports_detected": summary.get("reports_detected"), "total_files": summary.get("total_files")}, "H_B")
        enforce_plan(user, summary)
        hotel_context = {
            "hotel_nombre": user["hotel_name"],
            "hotel_tamano": user["hotel_size"] or "",
            "hotel_categoria": user["hotel_category"] or "",
            "hotel_ubicacion": user["hotel_location"] or "",
            "hotel_estrellas": user.get("hotel_stars") or 0,
            "hotel_ubicacion_destino": user.get("hotel_location_context") or "",
            "hotel_pms": user.get("hotel_pms") or "",
            "hotel_channel_manager": user.get("hotel_channel_manager") or "",
            "hotel_booking_engine": user.get("hotel_booking_engine") or "",
            "hotel_tech_other": user.get("hotel_tech_other") or "",
            "hotel_google_business_url": user.get("hotel_google_business_url") or "",
            "hotel_expedia_url": user.get("hotel_expedia_url") or "",
            "hotel_booking_url": user.get("hotel_booking_url") or "",
        }
        combined_business_context = business_context or ""
        if user["plan"] == "free":
            plan_for_model = "free_30"
        elif user["plan"] == "pro":
            plan_for_model = "pro_90"
        else:
            plan_for_model = "pro_180"
        _dbg("app.py:analyze", "before_call_openai", {}, "H_D")
        analysis = call_openai(summary, combined_business_context, hotel_context, plan_for_model)
        title = f"{summary['reports_detected']} reporte(s) · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        analysis_id, share_token = save_analysis(user["id"], title, user["plan"], summary, analysis, files)
        created_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")[:19].replace("T", " ")
        share_url = f"{public_share_base_url()}/s/{share_token}"
        # #region agent log
        _debug_log("app.py:analyze", "POST /analyze success", {"analysis_id": analysis_id}, "H4")
        _dbg("app.py:analyze", "success", {"analysis_id": analysis_id}, "H_D")
        # #endregion
        return JSONResponse({
            "ok": True,
            "analysis_id": analysis_id,
            "title": title,
            "created_at": created_at,
            "summary": summary,
            "analysis": analysis,
            "plan": user["plan"],
            "share_url": share_url,
        })
    except HTTPException as e:
        _dbg("app.py:analyze", "http_exception", {"detail": e.detail, "status_code": e.status_code}, "H_C")
        return JSONResponse({"ok": False, "error": e.detail}, status_code=e.status_code)
    except ValueError as e:
        _dbg("app.py:analyze", "value_error", {"message": str(e)}, "H_B")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        # #region agent log
        _debug_log("app.py:analyze", "POST /analyze exception", {"error": str(e)}, "H4")
        _dbg("app.py:analyze", "exception", {"exc_type": type(e).__name__, "exc_msg": str(e)}, "H_B")
        # #endregion
        err_msg = str(e).strip() if str(e) else "Error desconocido"
        return JSONResponse({"ok": False, "error": f"No se pudo completar el análisis. Intenta de nuevo más tarde. ({err_msg})"}, status_code=500)


def web_analysis_detail_json(request: Request, analysis_id: int):
    user = require_user(request)
    with db() as conn:
        row = conn.execute("SELECT * FROM analyses WHERE id = ? AND user_id = ?", (analysis_id, user["id"])).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Análisis no encontrado")
    stoken = row["share_token"] if row["share_token"] else None
    share_url = f"{public_share_base_url()}/s/{stoken}" if stoken else None
    return JSONResponse({
        "ok": True,
        "summary": json.loads(row["summary_json"]),
        "analysis": json.loads(row["analysis_json"]),
        "id": row["id"],
        "title": row["title"],
        "created_at": row["created_at"],
        "plan": row["plan_at_analysis"] or "free",
        "share_url": share_url,
    })


def web_analysis_share_ensure(request: Request, analysis_id: int):
    """Crea o devuelve el enlace público de solo lectura para un análisis (p. ej. análisis guardados antes de la migración)."""
    user = require_user(request)
    with db() as conn:
        row = conn.execute(
            "SELECT share_token FROM analyses WHERE id = ? AND user_id = ?",
            (analysis_id, user["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Análisis no encontrado")
        token = row["share_token"]
        if not token:
            token = secrets.token_urlsafe(24)
            conn.execute("UPDATE analyses SET share_token = ? WHERE id = ?", (token, analysis_id))
    return JSONResponse({"ok": True, "share_url": f"{public_share_base_url()}/s/{token}"})


async def web_analysis_share_email(request: Request, analysis_id: int, to_email: str):
    """Envía por SMTP el enlace público del análisis a un correo (requiere SMTP configurado en el servidor)."""
    user = require_user(request)
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASSWORD:
        raise HTTPException(
            status_code=503,
            detail="El envío por correo no está configurado en el servidor. Usa “Abrir en mi correo” o configura SMTP (.env).",
        )
    to_clean = (to_email or "").strip()
    if not looks_like_email(to_clean):
        raise HTTPException(status_code=400, detail="Correo no válido.")
    uid = int(user["id"])
    hotel_label = (user["hotel_name"] or "").strip() or "Hotel"
    with db() as conn:
        row = conn.execute(
            "SELECT share_token FROM analyses WHERE id = ? AND user_id = ?",
            (analysis_id, uid),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Análisis no encontrado")
        token = row["share_token"]
        if not token:
            token = secrets.token_urlsafe(24)
            conn.execute("UPDATE analyses SET share_token = ? WHERE id = ?", (token, analysis_id))
    share_url = f"{public_share_base_url()}/s/{token}"
    if not send_analysis_share_link_email(to_clean, share_url, hotel_label):
        raise HTTPException(status_code=500, detail="No se pudo enviar el correo. Intenta más tarde.")
    return JSONResponse({"ok": True, "message": "Correo enviado."})


def web_shared_analysis_page(request: Request, share_token: str):
    """Vista pública de solo lectura del análisis (quien tenga el enlace)."""
    with db() as conn:
        row = conn.execute(
            "SELECT title, summary_json, analysis_json, created_at FROM analyses WHERE share_token = ?",
            (share_token,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Enlace no válido o expirado.")
    summary = json.loads(row["summary_json"])
    analysis = json.loads(row["analysis_json"])
    created = (row["created_at"] or "")[:19].replace("T", " ")
    return templates.TemplateResponse(
        "share_public.html",
        {
            "request": request,
            "page_title": row["title"] or "Informe compartido",
            "created_at": created,
            "summary": summary,
            "analysis": analysis,
        },
    )


# Colores brandbook para PDF
_PDF_BRAND_TEXT = HexColor("#343434")
_PDF_BRAND_ACCENT = HexColor("#f07e07")
_PDF_MARGIN = 50
_PDF_FOOTER_H = 36


def _pdf_draw_footer(c: canvas.Canvas, width: float, height: float, hotel_name: str, report_date: str) -> None:
    """Pie de página: nombre del hotel y fecha en todas las páginas."""
    c.setFillColor(_PDF_BRAND_TEXT)
    c.setFont("Helvetica", 9)
    c.drawString(_PDF_MARGIN, 22, f"{hotel_name}  ·  {report_date}")
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.setStrokeColor(_PDF_BRAND_ACCENT)
    c.setLineWidth(0.5)
    c.line(_PDF_MARGIN, 30, width - _PDF_MARGIN, 30)
    c.setFillColor(_PDF_BRAND_TEXT)


def _pdf_draw_wrapped(
    c: canvas.Canvas, text: str, x: float, y: float, max_width: int, font: str = "Helvetica", size: int = 10, leading: int = 14
) -> float:
    """Dibuja texto con salto de línea; devuelve la y final."""
    c.setFont(font, size)
    lines: list[str] = []
    for raw_line in (text or "").split("\n"):
        current = ""
        for word in raw_line.split():
            test = (current + " " + word).strip()
            if stringWidth(test, font, size) > max_width:
                if current:
                    lines.append(current)
                current = word
            else:
                current = test
        if current:
            lines.append(current)
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    return y


def _pdf_build_analysis_pdf(
    c: canvas.Canvas, width: float, height: float, user: dict, row: dict, summary: dict, analysis: dict
) -> None:
    hotel_name = user.get("hotel_name") or "Hotel"
    plan = plan_label(user.get("plan") or "free")
    report_date = (row.get("created_at") or "")[:19].replace("T", " ")
    analysis_id = row.get("id", 0)

    # Pie de página en la primera hoja
    _pdf_draw_footer(c, width, height, hotel_name, report_date)

    def new_page(y: float, min_need: int = 120) -> float:
        if y < _PDF_FOOTER_H + min_need:
            c.showPage()
            _pdf_draw_footer(c, width, height, hotel_name, report_date)
            return height - _PDF_MARGIN - 20
        return y

    # ---------- Encabezado con branding ----------
    logo_path = BASE_DIR / "static" / "branding" / "dragonne-wordmark.png"
    if logo_path.exists():
        try:
            c.drawImage(str(logo_path), _PDF_MARGIN, height - 52, width=120, height=28)
        except Exception:
            c.setFillColor(_PDF_BRAND_ACCENT)
            c.setFont("Helvetica-Bold", 20)
            c.drawString(_PDF_MARGIN, height - 48, "DRAGONNÉ")
            c.setFillColor(_PDF_BRAND_TEXT)
    else:
        c.setFillColor(_PDF_BRAND_ACCENT)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(_PDF_MARGIN, height - 48, "DRAGONNÉ")
        c.setFillColor(_PDF_BRAND_TEXT)
    c.setFont("Helvetica", 11)
    c.drawString(_PDF_MARGIN, height - 68, "Análisis de revenue hotelero")
    c.setStrokeColor(_PDF_BRAND_ACCENT)
    c.setLineWidth(1.5)
    c.line(_PDF_MARGIN, height - 76, min(_PDF_MARGIN + 180, width - _PDF_MARGIN), height - 76)
    c.setFillColor(_PDF_BRAND_TEXT)

    # ---------- Bloque personalizado: Hotel y fecha ----------
    y = height - 100
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Hotel")
    c.setFillColor(_PDF_BRAND_TEXT)
    c.setFont("Helvetica", 10)
    c.drawString(_PDF_MARGIN + 32, y, hotel_name)
    y -= 14
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Fecha del reporte")
    c.setFillColor(_PDF_BRAND_TEXT)
    c.setFont("Helvetica", 10)
    c.drawString(_PDF_MARGIN + 70, y, report_date)
    y -= 14
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Plan")
    c.setFillColor(_PDF_BRAND_TEXT)
    c.setFont("Helvetica", 10)
    c.drawString(_PDF_MARGIN + 28, y, plan)
    y -= 28

    # ---------- Tabla resumen (KPIs como en el dashboard) ----------
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Resumen del análisis")
    y -= 6
    c.setStrokeColor(HexColor("#e0e0e0"))
    c.setLineWidth(0.5)
    tw = (width - 2 * _PDF_MARGIN) / 4
    th = 22
    kpis = [
        ("Archivos", str(summary.get("total_files", 0))),
        ("Reportes", str(summary.get("reports_detected", 0))),
        ("Días cubiertos", str(summary.get("overall_days_covered", 0))),
        ("Máx. rango", str(summary.get("max_days_covered", 0))),
    ]
    for i, (label, value) in enumerate(kpis):
        cx = _PDF_MARGIN + i * tw
        c.rect(cx, y - th, tw, th, stroke=1, fill=0)
        c.setFont("Helvetica", 8)
        c.setFillColor(HexColor("#808081"))
        c.drawString(cx + 6, y - th + 12, label)
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(_PDF_BRAND_TEXT)
        c.drawString(cx + 6, y - th + 2, value)
    c.setFillColor(_PDF_BRAND_TEXT)
    y -= th + 18

    # ---------- Resumen ejecutivo ----------
    y = new_page(y, 80)
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Resumen ejecutivo")
    y -= 6
    c.setStrokeColor(_PDF_BRAND_ACCENT)
    c.setLineWidth(1)
    c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 120, y - 2)
    c.setFillColor(_PDF_BRAND_TEXT)
    y -= 16
    resumen = analysis.get("resumen_ejecutivo", "")
    y = _pdf_draw_wrapped(c, resumen, _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
    y -= 14

    # ---------- Métricas clave ----------
    y = new_page(y, 100)
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Métricas clave")
    y -= 6
    c.setStrokeColor(_PDF_BRAND_ACCENT)
    c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 100, y - 2)
    c.setFillColor(_PDF_BRAND_TEXT)
    y -= 14
    c.setFont("Helvetica", 10)
    for item in analysis.get("metricas_clave", [])[:8]:
        nombre = item.get("nombre", "")
        valor = item.get("valor", "")
        lectura = item.get("lectura", "")
        y = _pdf_draw_wrapped(c, f"{nombre}: {valor}", _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
        if lectura:
            y = _pdf_draw_wrapped(c, lectura, _PDF_MARGIN + 10, y, int(width - 2 * _PDF_MARGIN - 10), size=9)
        y -= 6
        if y < _PDF_FOOTER_H + 60:
            y = new_page(y, 60)
            c.setFont("Helvetica", 10)
    y -= 8

    # ---------- Hallazgos prioritarios ----------
    y = new_page(y, 100)
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Hallazgos prioritarios")
    y -= 6
    c.setStrokeColor(_PDF_BRAND_ACCENT)
    c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 130, y - 2)
    c.setFillColor(_PDF_BRAND_TEXT)
    y -= 14
    for item in analysis.get("hallazgos_prioritarios", [])[:6]:
        titulo = item.get("titulo", "")
        detalle = item.get("detalle", "")
        impacto = item.get("impacto", "")
        prioridad = item.get("prioridad", "")
        y = _pdf_draw_wrapped(c, f"• {titulo} (Impacto: {impacto}, Prioridad: {prioridad})", _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
        if detalle:
            y = _pdf_draw_wrapped(c, detalle, _PDF_MARGIN + 12, y, int(width - 2 * _PDF_MARGIN - 12), size=9)
        y -= 8
        if y < _PDF_FOOTER_H + 50:
            y = new_page(y, 50)
    y -= 8

    # ---------- Oportunidades directo vs OTA ----------
    opps = analysis.get("oportunidades_directo_vs_ota", [])
    if opps:
        y = new_page(y, 80)
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(_PDF_BRAND_ACCENT)
        c.drawString(_PDF_MARGIN, y, "Oportunidades directo vs OTA")
        y -= 6
        c.setStrokeColor(_PDF_BRAND_ACCENT)
        c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 180, y - 2)
        c.setFillColor(_PDF_BRAND_TEXT)
        y -= 14
        c.setFont("Helvetica", 10)
        for s in opps[:6]:
            y = _pdf_draw_wrapped(c, f"• {s}", _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
            y -= 4
        y -= 8

    # ---------- Riesgos detectados ----------
    riesgos = analysis.get("riesgos_detectados", [])
    if riesgos:
        y = new_page(y, 80)
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(_PDF_BRAND_ACCENT)
        c.drawString(_PDF_MARGIN, y, "Riesgos detectados")
        y -= 6
        c.setStrokeColor(_PDF_BRAND_ACCENT)
        c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 120, y - 2)
        c.setFillColor(_PDF_BRAND_TEXT)
        y -= 14
        c.setFont("Helvetica", 10)
        for s in riesgos[:6]:
            y = _pdf_draw_wrapped(c, f"• {s}", _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
            y -= 4
        y -= 8

    # ---------- Recomendaciones accionables ----------
    y = new_page(y, 120)
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(_PDF_BRAND_ACCENT)
    c.drawString(_PDF_MARGIN, y, "Recomendaciones accionables")
    y -= 6
    c.setStrokeColor(_PDF_BRAND_ACCENT)
    c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 200, y - 2)
    c.setFillColor(_PDF_BRAND_TEXT)
    y -= 14
    c.setFont("Helvetica", 10)
    for item in analysis.get("recomendaciones_accionables", [])[:10]:
        accion = item.get("accion", "")
        por_que = item.get("por_que", "")
        urgencia = item.get("urgencia", "")
        y = _pdf_draw_wrapped(c, f"• {accion} (Urgencia: {urgencia})", _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
        if por_que:
            y = _pdf_draw_wrapped(c, por_que, _PDF_MARGIN + 12, y, int(width - 2 * _PDF_MARGIN - 12), size=9)
        y -= 6
        if y < _PDF_FOOTER_H + 50:
            y = new_page(y, 50)
            c.setFont("Helvetica", 10)
    y -= 8

    # ---------- Datos faltantes ----------
    faltantes = analysis.get("datos_faltantes", [])
    if faltantes:
        y = new_page(y, 60)
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(_PDF_BRAND_ACCENT)
        c.drawString(_PDF_MARGIN, y, "Datos faltantes")
        y -= 6
        c.setStrokeColor(_PDF_BRAND_ACCENT)
        c.line(_PDF_MARGIN, y - 2, _PDF_MARGIN + 100, y - 2)
        c.setFillColor(_PDF_BRAND_TEXT)
        y -= 14
        c.setFont("Helvetica", 10)
        for s in faltantes[:5]:
            y = _pdf_draw_wrapped(c, f"• {s}", _PDF_MARGIN, y, int(width - 2 * _PDF_MARGIN))
            y -= 4

    c.showPage()


def web_analysis_pdf_download(request: Request, analysis_id: int):
    user = require_user(request)
    with db() as conn:
        row = conn.execute(
            "SELECT * FROM analyses WHERE id = ? AND user_id = ?",
            (analysis_id, user["id"]),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Análisis no encontrado")

    summary = json.loads(row["summary_json"])
    analysis = json.loads(row["analysis_json"])
    user_dict = dict(user)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    _pdf_build_analysis_pdf(c, width, height, user_dict, dict(row), summary, analysis)
    c.save()
    buffer.seek(0)

    safe_hotel = re.sub(r"[^\w\s-]", "", (user_dict.get("hotel_name") or "informe"))
    safe_hotel = re.sub(r"[-\s]+", "-", safe_hotel).strip()[:40] or "informe"
    report_date = (row["created_at"] or "")[:10]
    filename = f"dragonne-informe-{safe_hotel}-{report_date}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


# ---------- API v1 (documentada en /docs y /redoc) ----------
@api_v1.get("/me", summary="Perfil del usuario asociado a la API key")
def api_me(user: sqlite3.Row = Depends(get_api_user)):
    """Devuelve el perfil del usuario autenticado por API key (solo campos no sensibles)."""
    return {
        "id": user["id"],
        "email": user["email"],
        "hotel_name": user["hotel_name"],
        "plan": user["plan"],
        "hotel_size": user["hotel_size"],
        "hotel_category": user["hotel_category"],
        "hotel_location": user["hotel_location"],
    }


@api_v1.post("/analyze", summary="Ejecutar análisis de reportes")
async def api_analyze(
    user: sqlite3.Row = Depends(get_api_user),
    business_context: str = Form(""),
    files: List[UploadFile] = File(...),
):
    """
    Sube uno o más reportes (CSV/Excel) y devuelve el análisis en JSON.
    Mismo límite de plan que en la web (días, archivos, número de análisis).
    """
    user = _user_row_as_dict(user)
    if not files:
        raise HTTPException(status_code=400, detail="Sube al menos un reporte.")
    try:
        summary = summarize_reports(files)
        enforce_plan(user, summary)
        hotel_context = {
            "hotel_nombre": user["hotel_name"],
            "hotel_tamano": user["hotel_size"] or "",
            "hotel_categoria": user["hotel_category"] or "",
            "hotel_ubicacion": user["hotel_location"] or "",
            "hotel_estrellas": user.get("hotel_stars") or 0,
            "hotel_ubicacion_destino": user.get("hotel_location_context") or "",
            "hotel_pms": user.get("hotel_pms") or "",
            "hotel_channel_manager": user.get("hotel_channel_manager") or "",
            "hotel_booking_engine": user.get("hotel_booking_engine") or "",
            "hotel_tech_other": user.get("hotel_tech_other") or "",
            "hotel_google_business_url": user.get("hotel_google_business_url") or "",
            "hotel_expedia_url": user.get("hotel_expedia_url") or "",
            "hotel_booking_url": user.get("hotel_booking_url") or "",
        }
        combined_business_context = business_context or ""
        if user["plan"] == "free":
            plan_for_model = "free_30"
        elif user["plan"] == "pro":
            plan_for_model = "pro_90"
        else:
            plan_for_model = "pro_180"
        analysis = call_openai(summary, combined_business_context, hotel_context, plan_for_model)
        title = f"{summary['reports_detected']} reporte(s) · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        analysis_id, share_token = save_analysis(user["id"], title, user["plan"], summary, analysis, files)
        return {
            "ok": True,
            "analysis_id": analysis_id,
            "title": title,
            "summary": summary,
            "analysis": analysis,
            "plan": user["plan"],
            "share_url": f"{public_share_base_url()}/s/{share_token}",
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="No se pudo completar el análisis. Intenta más tarde.")


@api_v1.get("/analyses", summary="Listar análisis del usuario")
def api_list_analyses(user: sqlite3.Row = Depends(get_api_user)):
    """Lista los análisis del usuario (últimos 50)."""
    with db() as conn:
        rows = conn.execute(
            "SELECT id, title, plan_at_analysis, file_count, days_covered, created_at FROM analyses WHERE user_id = ? ORDER BY created_at DESC LIMIT 50",
            (user["id"],),
        ).fetchall()
    return {
        "ok": True,
        "analyses": [
            {
                "id": row["id"],
                "title": row["title"],
                "plan_at_analysis": row["plan_at_analysis"],
                "file_count": row["file_count"],
                "days_covered": row["days_covered"],
                "created_at": row["created_at"],
            }
            for row in rows
        ],
    }


@api_v1.get("/analyses/{analysis_id}", summary="Obtener un análisis por ID")
def api_get_analysis(analysis_id: int, user: sqlite3.Row = Depends(get_api_user)):
    """Devuelve el JSON completo de un análisis (summary + analysis)."""
    with db() as conn:
        row = conn.execute("SELECT * FROM analyses WHERE id = ? AND user_id = ?", (analysis_id, user["id"])).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Análisis no encontrado")
    return {
        "ok": True,
        "id": row["id"],
        "title": row["title"],
        "created_at": row["created_at"],
        "summary": json.loads(row["summary_json"]),
        "analysis": json.loads(row["analysis_json"]),
    }


@api_v1.get("/analyses/{analysis_id}/pdf", summary="Descargar PDF del análisis")
def api_analysis_pdf(analysis_id: int, user: sqlite3.Row = Depends(get_api_user)):
    """Genera y devuelve el PDF del análisis (mismo formato que la web: branding, tabla resumen, hotel y fecha)."""
    with db() as conn:
        row = conn.execute("SELECT * FROM analyses WHERE id = ? AND user_id = ?", (analysis_id, user["id"])).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Análisis no encontrado")
    summary = json.loads(row["summary_json"])
    analysis = json.loads(row["analysis_json"])
    user_dict = dict(user)
    row_dict = dict(row)
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    _pdf_build_analysis_pdf(c, width, height, user_dict, row_dict, summary, analysis)
    c.save()
    buffer.seek(0)
    safe_hotel = re.sub(r"[^\w\s-]", "", (user_dict.get("hotel_name") or "informe"))
    safe_hotel = re.sub(r"[-\s]+", "-", safe_hotel).strip()[:40] or "informe"
    report_date = (row["created_at"] or "")[:10]
    filename = f"dragonne-informe-{safe_hotel}-{report_date}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


app.include_router(api_v1)

from routes.admin import router as admin_router
from routes.analysis import router as analysis_router
from routes.auth import router as auth_router
from routes.consulting import router as consulting_router
from routes.marketing import router as marketing_router

app.include_router(marketing_router)
app.include_router(auth_router)
app.include_router(consulting_router)
app.include_router(admin_router)
app.include_router(analysis_router)


@app.post("/billing/create-checkout-session")
def create_checkout_session(request: Request, billing_cycle: str = Form(...), plan_tier: str = Form("pro")):
    user = require_user(request)
    if plan_tier == "pro_plus":
        if not STRIPE_PRO_PLUS_PRICE_ID:
            raise HTTPException(status_code=500, detail="Pro+ no configurado (STRIPE_PRO_PLUS_PRICE_ID)")
        price_id = STRIPE_PRO_PLUS_PRICE_ID
    else:
        if billing_cycle not in {"monthly", "annual"}:
            raise HTTPException(status_code=400, detail="Ciclo inválido")
        price_id = STRIPE_MONTHLY_PRICE_ID if billing_cycle == "monthly" else STRIPE_ANNUAL_PRICE_ID
        if not price_id:
            raise HTTPException(status_code=500, detail="Falta configurar el price_id de Stripe (Pro)")
    customer_id = ensure_stripe_customer(user)
    payload = {
        "mode": "subscription",
        "customer": customer_id,
        "success_url": f"{APP_URL}/billing/success?session_id={{CHECKOUT_SESSION_ID}}",
        "cancel_url": f"{APP_URL}/app",
        "line_items[0][price]": price_id,
        "line_items[0][quantity]": 1,
        "allow_promotion_codes": "true",
        "metadata[user_id]": str(user["id"]),
        "metadata[billing_cycle]": billing_cycle if plan_tier != "pro_plus" else "monthly",
        "metadata[plan_tier]": plan_tier,
    }
    if TRIAL_DAYS > 0 and plan_tier != "pro_plus":
        payload["subscription_data[trial_period_days]"] = TRIAL_DAYS
    session = stripe_request("POST", "/v1/checkout/sessions", payload)
    return JSONResponse({"ok": True, "url": session["url"]})


@app.get("/billing/success", response_class=HTMLResponse)
def billing_success(request: Request):
    user = require_user(request)
    return templates.TemplateResponse("billing_success.html", {"request": request, "user": user})


@app.post("/billing/create-portal-session")
def create_portal_session(request: Request):
    user = require_user(request)
    if not user["stripe_customer_id"]:
        raise HTTPException(status_code=400, detail="No hay cliente de Stripe asociado todavía.")
    session = stripe_request("POST", "/v1/billing_portal/sessions", {
        "customer": user["stripe_customer_id"],
        "return_url": f"{APP_URL}/app",
    })
    return JSONResponse({"ok": True, "url": session["url"]})


@app.post("/billing/webhook")
async def billing_webhook(request: Request):
    payload = await request.body()
    sig = request.headers.get("Stripe-Signature", "")
    if not STRIPE_WEBHOOK_SECRET:
        raise HTTPException(status_code=500, detail="Webhook no configurado")
    header_parts = dict(item.split("=", 1) for item in sig.split(",") if "=" in item)
    timestamp = header_parts.get("t", "")
    signature = header_parts.get("v1", "")
    signed_payload = f"{timestamp}.{payload.decode()}".encode()
    expected = hmac.new(STRIPE_WEBHOOK_SECRET.encode(), signed_payload, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, signature):
        raise HTTPException(status_code=400, detail="Firma inválida")
    event = json.loads(payload.decode())
    event_id = event.get("id")
    event_type = event.get("type")
    with db() as conn:
        exists = conn.execute("SELECT id FROM billing_events WHERE stripe_event_id = ?", (event_id,)).fetchone()
        if exists:
            return Response(status_code=200)
        conn.execute(
            "INSERT INTO billing_events (stripe_event_id, event_type, payload, created_at) VALUES (?, ?, ?, ?)",
            (event_id, event_type or "unknown", payload.decode(), now_iso()),
        )
    obj = event.get("data", {}).get("object", {})
    if event_type in {"checkout.session.completed"}:
        customer_id = obj.get("customer")
        subscription_id = obj.get("subscription")
        if customer_id:
            sync_user_from_stripe_customer(customer_id, subscription_id, "active")
    elif event_type in {"customer.subscription.created", "customer.subscription.updated"}:
        customer_id = obj.get("customer")
        status = obj.get("status", "")
        subscription_id = obj.get("id")
        if customer_id:
            sync_user_from_stripe_customer(customer_id, subscription_id, status)
    elif event_type in {"customer.subscription.deleted"}:
        customer_id = obj.get("customer")
        if customer_id:
            sync_user_from_stripe_customer(customer_id, None, "canceled")
    return Response(status_code=200)


@app.get("/health")
def health():
    """Solo estado básico; no exponer configuración interna."""
    return {"ok": True, "app": APP_NAME}


@app.get("/health/config")
def health_config():
    """Indica si las variables de entorno críticas están definidas (solo sí/no, sin valores)."""
    return {
        "ok": True,
        "openai_configured": bool(OPENAI_API_KEY and OPENAI_API_KEY.strip()),
        "stripe_configured": bool(STRIPE_SECRET_KEY and STRIPE_SECRET_KEY.strip()),
        "stripe_webhook_configured": bool(STRIPE_WEBHOOK_SECRET and STRIPE_WEBHOOK_SECRET.strip()),
        "stripe_pro_price_configured": bool(STRIPE_MONTHLY_PRICE_ID and STRIPE_MONTHLY_PRICE_ID.strip()),
        "stripe_pro_plus_price_configured": bool(STRIPE_PRO_PLUS_PRICE_ID and STRIPE_PRO_PLUS_PRICE_ID.strip()),
    }
